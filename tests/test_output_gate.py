from __future__ import annotations

import csv
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from sandbox_tool.output_gate import run_output_gate

FIXTURE_ROOT = Path(__file__).parent / "security_fixtures"


class OutputGateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.raw_root = self.root / "raw_outputs"
        self.clean_root = self.root / "clean_exports"
        self.quarantine_root = self.root / "quarantine"
        self.log_root = self.root / "gate_logs"
        for path in (self.raw_root, self.clean_root, self.quarantine_root, self.log_root):
            path.mkdir(parents=True, exist_ok=True)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def gate(self, *artifacts: str, action: str = "reject") -> dict:
        return run_output_gate(
            raw_root=self.raw_root,
            clean_root=self.clean_root,
            quarantine_root=self.quarantine_root,
            log_root=self.log_root,
            artifacts=list(artifacts),
            run_id="unit-test",
            xlsx_dangerous_formula_action=action,
        )

    def test_markdown_passes(self) -> None:
        (self.raw_root / "report.md").write_text("# Report\n\nNo active content.\n", encoding="utf-8")

        manifest = self.gate("/outputs/report.md")

        self.assertEqual(manifest["overall_status"], "pass")
        self.assertTrue((self.clean_root / "report.md").exists())
        self.assertEqual(manifest["artifacts"][0]["status"], "pass")

    def test_markdown_security_good_fixture_passes(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "good.md", self.raw_root / "good.md")

        manifest = self.gate("/outputs/good.md")

        self.assertEqual(manifest["overall_status"], "pass")
        self.assertTrue((self.clean_root / "good.md").exists())

    def test_html_security_good_fixture_passes_with_csp(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "good.html", self.raw_root / "good.html")

        manifest = self.gate("/outputs/good.html")

        self.assertEqual(manifest["overall_status"], "pass")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "sanitized")
        self.assertIn("injected_csp", artifact["actions"])
        self.assertIn(
            "Content-Security-Policy",
            (self.clean_root / "good.html").read_text(encoding="utf-8"),
        )

    def test_html_script_is_rejected_and_quarantined(self) -> None:
        (self.raw_root / "index.html").write_text("<html><script>alert(1)</script></html>", encoding="utf-8")

        manifest = self.gate("/outputs/index.html")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "script_tag")
        self.assertTrue((self.quarantine_root / "index.html").exists())

    def test_csv_formula_cells_are_escaped(self) -> None:
        (self.raw_root / "data.csv").write_text("name,value\nsafe,10\nbad,=1+1\n", encoding="utf-8")

        manifest = self.gate("/outputs/data.csv")

        self.assertEqual(manifest["overall_status"], "pass")
        self.assertEqual(manifest["artifacts"][0]["status"], "sanitized")
        with (self.clean_root / "data.csv").open(newline="", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        self.assertEqual(rows[2][1], "'=1+1")

    def test_csv_negative_numbers_are_not_escaped(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "good.csv", self.raw_root / "good.csv")

        manifest = self.gate("/outputs/good.csv")

        self.assertEqual(manifest["overall_status"], "pass")
        with (self.clean_root / "good.csv").open(newline="", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        values = {row[0]: row[1] for row in rows[1:]}
        self.assertEqual(values["refund"], "-100")
        self.assertEqual(values["adjustment"], "+25.5")
        self.assertEqual(values["scientific"], "-1.25e3")

    def test_csv_security_fixture_escapes_formulas_but_preserves_numbers(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "evil.csv", self.raw_root / "evil.csv")

        manifest = self.gate("/outputs/evil.csv")

        self.assertEqual(manifest["overall_status"], "pass")
        self.assertEqual(manifest["artifacts"][0]["status"], "sanitized")
        with (self.clean_root / "evil.csv").open(newline="", encoding="utf-8") as handle:
            rows = list(csv.reader(handle))
        values = {row[0]: row[1] for row in rows[1:]}
        self.assertEqual(values["formula"], "'=WEBSERVICE(\"https://example.com\")")
        self.assertEqual(values["plus"], "'+cmd|")
        self.assertEqual(values["at"], "'@SUM(1,1)")
        self.assertEqual(values["negative_formula"], "'-SUM(1,1)")
        self.assertEqual(values["normal_negative"], "-100")
        self.assertEqual(values["normal_positive"], "+100")

    def test_csv_disguised_zip_is_rejected(self) -> None:
        (self.raw_root / "evil.csv").write_bytes(b"PK\x03\x04not actually csv")

        manifest = self.gate("/outputs/evil.csv")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "magic_extension_mismatch")
        self.assertTrue((self.quarantine_root / "evil.csv").exists())

    def test_markdown_disguised_html_document_is_rejected(self) -> None:
        (self.raw_root / "report.md").write_text(
            "<!doctype html><html><body>not markdown</body></html>",
            encoding="utf-8",
        )

        manifest = self.gate("/outputs/report.md")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "declared_format_mismatch")

    def test_html_without_document_marker_is_rejected(self) -> None:
        (self.raw_root / "fragment.html").write_text("<p>fragment only</p>", encoding="utf-8")

        manifest = self.gate("/outputs/fragment.html")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "html_document_marker_missing")

    def test_html_security_fixture_is_rejected(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "evil.html", self.raw_root / "evil.html")

        manifest = self.gate("/outputs/evil.html")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertIn(artifact["findings"][0]["code"], {"script_tag", "event_handler"})

    def test_markdown_security_fixture_is_rejected(self) -> None:
        shutil.copy2(FIXTURE_ROOT / "evil.md", self.raw_root / "evil.md")

        manifest = self.gate("/outputs/evil.md")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "script_tag")

    def test_xlsx_preserves_safe_formulas(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = 1
        sheet["A2"] = 2
        sheet["A3"] = "=SUM(A1:A2)"
        workbook.save(self.raw_root / "safe.xlsx")
        workbook.close()

        manifest = self.gate("/outputs/safe.xlsx")

        self.assertEqual(manifest["overall_status"], "pass")
        clean_workbook = openpyxl.load_workbook(self.clean_root / "safe.xlsx", data_only=False)
        try:
            self.assertEqual(clean_workbook.active["A3"].value, "=SUM(A1:A2)")
        finally:
            clean_workbook.close()

    def test_xlsx_dangerous_formula_is_rejected(self) -> None:
        workbook = openpyxl.Workbook()
        workbook.active["A1"] = '=WEBSERVICE("https://example.com")'
        workbook.save(self.raw_root / "danger.xlsx")
        workbook.close()

        manifest = self.gate("/outputs/danger.xlsx")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "xlsx_dangerous_formula_function")
        self.assertTrue((self.quarantine_root / "danger.xlsx").exists())

    def test_xlsx_text_file_is_rejected(self) -> None:
        (self.raw_root / "evil.xlsx").write_text("not an OOXML workbook", encoding="utf-8")

        manifest = self.gate("/outputs/evil.xlsx")

        self.assertEqual(manifest["overall_status"], "fail")
        artifact = manifest["artifacts"][0]
        self.assertEqual(artifact["status"], "rejected")
        self.assertEqual(artifact["findings"][0]["code"], "magic_extension_mismatch")
        self.assertTrue((self.quarantine_root / "evil.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
