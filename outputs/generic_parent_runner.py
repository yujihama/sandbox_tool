from __future__ import annotations

import argparse
import base64
import csv
import hashlib
from io import BytesIO
import json
import math
import mimetypes
import os
import posixpath
import re
import shutil
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import numpy as np
from dotenv import load_dotenv
from langchain.agents import create_agent
from langchain.agents.middleware import AgentMiddleware, AgentState, ModelRequest
from langchain.agents.middleware.types import ModelResponse
from langchain.tools import tool
from langchain_core.messages import SystemMessage, ToolMessage
from openai import OpenAI
from typing_extensions import NotRequired

from deepagents import create_deep_agent


THIS_FILE = Path(__file__).resolve()
ROOT = THIS_FILE.parents[1]
WORK_DIR = ROOT / "work"
OUTPUTS_ROOT = ROOT / "outputs"

for import_dir in (ROOT, THIS_FILE.parent, WORK_DIR):
    if str(import_dir) not in sys.path:
        sys.path.insert(0, str(import_dir))

from podman_sandbox_backend import PodmanRunLimits, PodmanSandboxBackend  # noqa: E402
from sandbox_tool.controller_sandbox_backend import ControllerSandboxBackend  # noqa: E402
from sandbox_tool.output_gate import (  # noqa: E402
    ALLOWED_EXTENSIONS,
    run_output_gate as run_output_gate_artifacts,
    sha256_file,
)
from sandbox_tool.houjin_bangou import (  # noqa: E402
    HoujinBangouSearchPolicy,
    run_houjin_bangou_search,
)
from sandbox_tool.egress_proxy import create_egress_token  # noqa: E402
from sandbox_tool.site_crawler import (  # noqa: E402
    CrawlPolicy,
    LinkExtractPolicy,
    extract_links_from_listing,
    is_private_or_local_host,
    list_crawls as list_site_crawl_runs,
    normalize_domain,
    read_crawled_page as read_site_crawl_page,
    run_site_crawl,
    run_url_crawl,
    search_crawl as search_site_crawl_index,
)


@dataclass
class InputMapping:
    host_path: Path
    sandbox_path: str
    staged_path: Path


@dataclass
class DeepAgentProfile:
    id: str
    tool_name: str
    description: str
    toolsets: list[str] = field(default_factory=list)
    expose_to_parent: bool = True
    input_access: str = "all"
    system_prompt: str = ""
    skill_source_specs: list[str] = field(default_factory=list)
    skill_sources: list[str] = field(default_factory=list)
    include_global_skills: bool = False
    image: str = ""
    deep_model: str = ""
    vision_model: str = ""
    deep_recursion_limit: int | None = None
    max_review_rounds: int | None = None
    graceful_finalize: dict[str, int] = field(default_factory=dict)
    result_mode: str = "artifact"
    self_check_policy: str = "script"
    source_path: str = ""


@dataclass
class RunnerConfig:
    prompt: str
    run_root: Path
    output_dir: Path
    clean_export_dir: Path
    quarantine_dir: Path
    gate_log_dir: Path
    runner_log_dir: Path
    input_dir: Path
    workspace_dir: Path
    input_mappings: list[InputMapping]
    expected_artifacts: list[str]
    skill_sources: list[str]
    image: str
    wsl_distro: str
    parent_model: str
    deep_model: str
    vision_model: str
    parent_recursion_limit: int
    deep_recursion_limit: int
    max_review_rounds: int
    host_os: str
    podman_bin: str
    wsl_use_sudo: bool
    selinux_relabel: bool
    clear_output: bool
    keep_staged_input: bool
    allow_raw_parent_inspection: bool
    sandbox_backend: str
    sandbox_controller_url: str
    sandbox_controller_token: str
    egress_proxy_url: str
    egress_proxy_signing_secret: str
    xlsx_dangerous_formula_action: str
    tile_small_images_for_vision: bool = False
    vision_tile_max_side: int = 128
    vision_tile_grid: int = 5
    deep_agent_profiles: list[DeepAgentProfile] = field(default_factory=list)


CONFIG: RunnerConfig | None = None
DEEP_AGENT_TRACE: list[dict[str, Any]] = []
DEEP_AGENT_EVALUATIONS: list[dict[str, Any]] = []
DEEP_REVIEW_REQUESTS: list[dict[str, Any]] = []
ALLOWED_EXPORT_EXTENSIONS_TEXT = ", ".join(sorted(ALLOWED_EXTENSIONS))
PLAYWRIGHT_DEFAULT_MIN_ACTION_DELAY_MS = 1000
PLAYWRIGHT_MAX_MIN_ACTION_DELAY_MS = 10000
PLAYWRIGHT_DEFAULT_RATE_LIMIT_BACKOFF_MS = 5000
PLAYWRIGHT_MAX_RATE_LIMIT_BACKOFF_MS = 60000
PLAYWRIGHT_DOMAIN_LAST_FINISH: dict[str, float] = {}
ACTIVE_DEEP_AGENT_TOOLSETS: set[str] = set()
ACTIVE_DEEP_AGENT_VISION_MODEL = ""
VISION_MIN_MARGIN_PX = 100
SKILL_VIRTUAL_PATH_RE = re.compile(
    r"(?P<path>/input/(?:skills|profile-skills)/(?P<skill>[^/\\\s\"'`]+)(?:/[^\\\s\"'`]+)*)"
)

TOOLSET_TOOL_NAMES: dict[str, list[str]] = {
    "review": ["request_parent_review"],
    "file_read": ["read_sandbox_file"],
    "image_inspect": ["inspect_sandbox_image"],
    "site_crawl": [
        "crawl_allowed_site",
        "extract_allowed_site_links",
        "crawl_allowed_urls",
        "search_site_crawl",
        "read_crawled_page",
        "list_site_crawls",
    ],
    "browser": ["run_playwright_task"],
}
TOOLSET_DESCRIPTIONS: dict[str, str] = {
    "review": "request parent review for gated artifacts",
    "file_read": "type-aware reads of /input and /outputs files",
    "image_inspect": "direct image vision inspection",
    "site_crawl": "controlled allowlisted site crawling and crawl-index reads",
    "browser": "deterministic Playwright browser interaction",
}
ALLOWED_TOOLSETS = set(TOOLSET_TOOL_NAMES)
FINALIZE_TOOL_ALLOWLIST = {
    "request_parent_review",
    "write_file",
    "edit_file",
    "execute",
}
STRICT_FINALIZE_TOOL_ALLOWLIST = {
    "request_parent_review",
}
REPAIR_FINALIZE_TOOL_ALLOWLIST = {
    "write_file",
    "edit_file",
    "execute",
}
GRACEFUL_FINALIZE_CONFIG_FIELDS = {
    "warning_model_calls",
    "finalize_model_calls",
    "warning_tool_calls",
    "finalize_tool_calls",
    "warning_message_count",
    "finalize_message_count",
    "strict_finalize_after_model_calls",
}


class GracefulFinalizeState(AgentState):
    graceful_finalize_model_calls: NotRequired[int]
    graceful_finalize_mode: NotRequired[str]


class GracefulFinalizeMiddleware(AgentMiddleware[GracefulFinalizeState]):
    """Shift a Deep Agent from exploration to artifact finalization before hard limits."""

    state_schema = GracefulFinalizeState

    def __init__(
        self,
        *,
        profile_id: str,
        expected_artifacts: list[str],
        warning_model_calls: int,
        finalize_model_calls: int,
        warning_tool_calls: int,
        finalize_tool_calls: int,
        warning_message_count: int,
        finalize_message_count: int,
        strict_finalize_after_model_calls: int = 2,
        finalize_tool_allowlist: set[str] | None = None,
        result_mode: str = "artifact",
        self_check_policy: str = "script",
    ) -> None:
        self.profile_id = profile_id
        self.expected_artifacts = expected_artifacts
        self.result_mode = result_mode
        self.self_check_policy = self_check_policy
        self.warning_model_calls = max(1, warning_model_calls)
        self.finalize_model_calls = max(self.warning_model_calls + 1, finalize_model_calls)
        self.warning_tool_calls = max(1, warning_tool_calls)
        self.finalize_tool_calls = max(self.warning_tool_calls + 1, finalize_tool_calls)
        self.warning_message_count = max(1, warning_message_count)
        self.finalize_message_count = max(
            self.warning_message_count + 1,
            finalize_message_count,
        )
        self.strict_finalize_after_model_calls = max(1, strict_finalize_after_model_calls)
        self.finalize_tool_allowlist = finalize_tool_allowlist or FINALIZE_TOOL_ALLOWLIST
        self.strict_finalize_tool_allowlist = STRICT_FINALIZE_TOOL_ALLOWLIST
        self.tool_calls = 0
        self.finalize_model_calls_seen = 0

    def before_model(
        self, state: GracefulFinalizeState, runtime: Any
    ) -> dict[str, Any] | None:
        count = state.get("graceful_finalize_model_calls", 0) + 1
        mode = self._mode_for_state(count, len(state.get("messages", [])), self.tool_calls)
        return {
            "graceful_finalize_model_calls": count,
            "graceful_finalize_mode": mode,
        }

    def wrap_model_call(
        self,
        request: ModelRequest,
        handler: Callable[[ModelRequest], ModelResponse],
    ) -> ModelResponse:
        count = int(request.state.get("graceful_finalize_model_calls", 0))
        message_count = len(request.state.get("messages", []))
        mode = self._mode_for_state(count, message_count, self.tool_calls)

        if mode == "finalize":
            self.finalize_model_calls_seen += 1
            return handler(
                request.override(
                    tools=self.filter_finalize_tools(
                        request.tools,
                        allowlist=self.current_finalize_tool_allowlist(),
                    ),
                    system_message=self._append_system_instruction(
                        request.system_message,
                        self.finalize_instruction(count, message_count),
                    ),
                )
            )

        if mode == "warning":
            return handler(
                request.override(
                    system_message=self._append_system_instruction(
                        request.system_message,
                        self.warning_instruction(count, message_count),
                    )
                )
            )

        return handler(request)

    def wrap_tool_call(
        self,
        request: Any,
        handler: Callable[[Any], Any],
    ) -> Any:
        self.tool_calls += 1
        tool_name = self._tool_name(getattr(request, "tool", None))
        if not tool_name:
            tool_name = self._tool_name(getattr(request, "tool_call", {}))
        mode = self._mode_for_state(
            int(request.state.get("graceful_finalize_model_calls", 0)),
            len(request.state.get("messages", [])),
            self.tool_calls,
        )
        if mode == "finalize" and tool_name not in self.current_finalize_tool_allowlist():
            tool_call = getattr(request, "tool_call", {}) or {}
            tool_call_id = str(tool_call.get("id") or tool_call.get("tool_call_id") or "")
            return ToolMessage(
                content=json.dumps(
                    {
                        "ok": False,
                        "error": "graceful_finalize_blocked_tool",
                        "tool": tool_name,
                        "message": (
                            "Execution budget is near the hard recursion limit. "
                            "Exploration tools are disabled; finalize with the "
                            "current profile's completion contract."
                        ),
                    },
                    ensure_ascii=False,
                ),
                tool_call_id=tool_call_id,
                name=tool_name or None,
            )
        return handler(request)

    def _mode_for_state(
        self,
        model_calls: int,
        message_count: int,
        tool_calls: int,
    ) -> str:
        if (
            model_calls >= self.finalize_model_calls
            or tool_calls >= self.finalize_tool_calls
            or message_count >= self.finalize_message_count
        ):
            return "finalize"
        if (
            model_calls >= self.warning_model_calls
            or tool_calls >= self.warning_tool_calls
            or message_count >= self.warning_message_count
        ):
            return "warning"
        return "normal"

    def _append_system_instruction(
        self,
        existing: SystemMessage | str | None,
        instruction: str,
    ) -> SystemMessage:
        if existing is None:
            return SystemMessage(content=instruction)
        if isinstance(existing, SystemMessage):
            base = existing.content
            if isinstance(base, list):
                base_text = "\n".join(str(part) for part in base)
            else:
                base_text = str(base)
            return SystemMessage(content=base_text + "\n\n" + instruction)
        return SystemMessage(content=str(existing) + "\n\n" + instruction)

    def _tool_name(self, item: Any) -> str:
        if isinstance(item, dict):
            if isinstance(item.get("function"), dict) and item["function"].get("name"):
                return str(item["function"]["name"])
            if item.get("name"):
                return str(item["name"])
        return str(getattr(item, "name", None) or getattr(item, "__name__", ""))

    def current_finalize_tool_allowlist(self) -> set[str]:
        if self.finalize_model_calls_seen >= self.strict_finalize_after_model_calls:
            if self.review_artifacts_look_ready():
                return self.strict_finalize_tool_allowlist
            return REPAIR_FINALIZE_TOOL_ALLOWLIST
        return self.finalize_tool_allowlist

    def review_artifacts_look_ready(self) -> bool:
        if self.result_mode != "artifact":
            return False
        try:
            for artifact in self.expected_artifacts:
                path = resolve_sandbox_path(artifact)
                if not path.is_file() or path.stat().st_size == 0:
                    return False
            if self.self_check_policy not in {"script", "checklist"}:
                return True
            plan_path = resolve_sandbox_path("/outputs/subtasks/self_check_plan.md")
            report_path = resolve_sandbox_path("/outputs/subtasks/self_check_report.md")
            if not plan_path.is_file() or not report_path.is_file():
                return False
            plan_text = plan_path.read_text(encoding="utf-8", errors="replace")
            report_text = report_path.read_text(encoding="utf-8", errors="replace")
            if review_markdown_blockers(plan_text, "self_check_plan.md"):
                return False
            report_findings = review_markdown_blockers(
                report_text,
                "self_check_report.md",
            )
            if report_findings:
                return False
            if self.self_check_policy == "script":
                lower = report_text.lower()
                if any(
                    marker in lower
                    for marker in (
                        "pending",
                        "not yet executed",
                        "not executed",
                        "未実行",
                    )
                ):
                    return False
                if not re.search(r"\b(?:pass|passed|self_check_pass)\b", lower):
                    return False
            return True
        except Exception:
            return False

    def filter_finalize_tools(
        self,
        tools: list[Any],
        *,
        allowlist: set[str] | None = None,
    ) -> list[Any]:
        effective_allowlist = allowlist or self.finalize_tool_allowlist
        return [
            item
            for item in tools
            if self._tool_name(item) in effective_allowlist
        ]

    def warning_instruction(self, model_calls: int, message_count: int) -> str:
        return (
            "[Budget warning]\n"
            f"Profile `{self.profile_id}` is approaching its execution budget "
            f"(model_calls={model_calls}, tool_calls={self.tool_calls}, "
            f"messages={message_count}). Stop broad "
            "exploration now: do not start new crawls/browser searches unless a "
            "specific required URL is already identified. Prefer existing collected "
            "evidence, narrow candidates, and move toward the profile's final "
            "answer/review contract."
        )

    def finalize_instruction(self, model_calls: int, message_count: int) -> str:
        if self.result_mode == "inline":
            return (
                "[Graceful finalize mode]\n"
                f"Profile `{self.profile_id}` is near its execution budget "
                f"(model_calls={model_calls}, tool_calls={self.tool_calls}, "
                f"messages={message_count}). Stop exploration and return the "
                "best supported inline answer now.\n\n"
                "Rules:\n"
                "- Do not perform new crawling, browser actions, broad searches, "
                "or new candidate discovery.\n"
                "- Do not create final reviewed artifacts and do not call "
                "request_parent_review.\n"
                "- Use only already collected evidence and files currently "
                "available in /outputs or /input.\n"
                "- Include a concise self-check checklist, uncertainty, and "
                "limitations directly in the final answer."
            )
        expected_text = "\n".join(f"- {path}" for path in self.expected_artifacts)
        review_artifacts = [*self.expected_artifacts]
        if self.self_check_policy in {"script", "checklist"}:
            review_artifacts.extend(
                [
                    "/outputs/subtasks/self_check_plan.md",
                    "/outputs/subtasks/self_check_report.md",
                ]
            )
        review_artifacts = list(dict.fromkeys(review_artifacts))
        review_text = "\n".join(f"- {path}" for path in review_artifacts)
        self_check_text = (
            "Create or update `/outputs/subtasks/self_check_plan.md` and "
            "`/outputs/subtasks/self_check_report.md`. Run a small self-check with "
            "`execute` if possible; if not, write the limitation in the report.\n\n"
            if self.self_check_policy == "script"
            else (
                "Create or update `/outputs/subtasks/self_check_plan.md` and "
                "`/outputs/subtasks/self_check_report.md` with a concise checklist "
                "self-check. Do not create or execute a separate self-check script.\n\n"
                if self.self_check_policy == "checklist"
                else "No self-check artifacts are required for this profile.\n\n"
            )
        )
        return (
            "[Graceful finalize mode]\n"
            f"Profile `{self.profile_id}` is near its execution budget "
            f"(model_calls={model_calls}, tool_calls={self.tool_calls}, "
            f"messages={message_count}). You must stop "
            "exploration and finalize now.\n\n"
            "Rules:\n"
            "- Do not perform new crawling, browser actions, broad searches, or new "
            "candidate discovery.\n"
            "- Use only already collected evidence and files currently available in "
            "/outputs or /input.\n"
            "- If fewer records than requested are supported, produce the supported "
            "partial result and clearly state the limitation.\n"
            "- If an expected artifact is missing, create a minimal honest artifact "
            "in the requested format rather than continuing exploration.\n\n"
            "Required expected artifacts:\n"
            f"{expected_text}\n\n"
            f"{self_check_text}"
            "When documenting active-content checks in Markdown, escape raw HTML "
            "tag examples such as `&lt;script&gt;`; do not write literal script "
            "tags in Markdown review artifacts because the output gate rejects "
            "active-content strings. Do not submit a placeholder or PENDING "
            "self-check report; request_parent_review will reject it. If a "
            "script self-check is required, execute the script and update the "
            "report with a final pass/fail result before requesting review.\n\n"
            "Then call request_parent_review with exactly these review artifacts:\n"
            f"{review_text}\n\n"
            "After request_parent_review returns, stop."
        )


def load_env_local(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def safe_tool_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
    name = name.strip("_-")
    if not name:
        name = "deep_agent"
    if not re.match(r"^[A-Za-z_]", name):
        name = "run_" + name
    return name[:64]


def as_string_list(value: Any, field_name: str) -> list[str]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError(f"{field_name} must be a list of strings.")
    result: list[str] = []
    for item in value:
        if not isinstance(item, str):
            raise ValueError(f"{field_name} must be a list of strings.")
        if item.strip():
            result.append(item.strip())
    return result


def validate_profile_toolsets(value: Any, profile_path: Path) -> list[str]:
    toolsets = as_string_list(value, "toolsets")
    if not toolsets:
        raise ValueError(f"Deep Agent profile must declare non-empty toolsets: {profile_path}")
    unknown = sorted(set(toolsets) - ALLOWED_TOOLSETS)
    if unknown:
        raise ValueError(
            f"Unknown toolsets in {profile_path}: {', '.join(unknown)}. "
            f"Allowed toolsets: {', '.join(sorted(ALLOWED_TOOLSETS))}"
        )
    if "review" not in toolsets:
        raise ValueError(f"Deep Agent profile must include the review toolset: {profile_path}")
    deduped: list[str] = []
    for toolset in toolsets:
        if toolset not in deduped:
            deduped.append(toolset)
    return deduped


def optional_bool(value: Any, field_name: str, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    raise ValueError(f"{field_name} must be a boolean when specified.")


def normalize_profile_input_access(value: Any, profile_path: Path) -> str:
    if value is None or value == "":
        return "all"
    access = str(value).strip().lower().replace("-", "_")
    if access not in {"all", "skills_only", "none"}:
        raise ValueError(
            f"Deep Agent profile input_access must be all, skills_only, or none: {profile_path}"
        )
    return access


def normalize_profile_result_mode(value: Any, profile_path: Path) -> str:
    if value is None or value == "":
        return "artifact"
    mode = str(value).strip().lower().replace("-", "_")
    if mode not in {"artifact", "inline"}:
        raise ValueError(
            f"Deep Agent profile result_mode must be artifact or inline: {profile_path}"
        )
    return mode


def normalize_profile_self_check_policy(value: Any, profile_path: Path) -> str:
    if value is None or value == "":
        return "script"
    policy = str(value).strip().lower().replace("-", "_")
    if policy not in {"script", "checklist", "none"}:
        raise ValueError(
            "Deep Agent profile self_check_policy must be script, checklist, "
            f"or none: {profile_path}"
        )
    return policy


def normalize_profile_graceful_finalize(value: Any, profile_path: Path) -> dict[str, int]:
    if value is None or value == "":
        return {}
    if not isinstance(value, dict):
        raise ValueError(f"graceful_finalize must be an object: {profile_path}")
    normalized: dict[str, int] = {}
    unknown = sorted(set(value) - GRACEFUL_FINALIZE_CONFIG_FIELDS)
    if unknown:
        raise ValueError(
            "Unsupported graceful_finalize field(s) "
            f"{', '.join(unknown)} in {profile_path}"
        )
    for key in GRACEFUL_FINALIZE_CONFIG_FIELDS:
        raw = value.get(key)
        if raw is None or raw == "":
            continue
        parsed = int(raw)
        if parsed <= 0:
            raise ValueError(f"graceful_finalize.{key} must be positive: {profile_path}")
        normalized[key] = parsed
    return normalized


def read_profile_document(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    text = path.read_text(encoding="utf-8")
    if suffix == ".json":
        data = json.loads(text)
    elif suffix in {".yaml", ".yml"}:
        import yaml

        data = yaml.safe_load(text)
    else:
        raise ValueError(f"Unsupported Deep Agent profile extension: {path}")
    if not isinstance(data, dict):
        raise ValueError(f"Deep Agent profile must be an object: {path}")
    return data


def load_deep_agent_profile(path: str | Path) -> DeepAgentProfile:
    profile_path = Path(path).expanduser().resolve()
    data = read_profile_document(profile_path)
    profile_id = str(data.get("id") or profile_path.stem).strip()
    if not profile_id:
        raise ValueError(f"Deep Agent profile id is empty: {profile_path}")
    tool_name = safe_tool_name(str(data.get("tool_name") or f"run_{profile_id}_agent"))
    description = str(
        data.get("description")
        or f"Run the {profile_id} Deep Agent profile in the sandbox."
    ).strip()
    system_prompt = str(data.get("system_prompt") or "").strip()
    system_prompt_file = data.get("system_prompt_file")
    if system_prompt_file:
        prompt_path = Path(str(system_prompt_file)).expanduser()
        if not prompt_path.is_absolute():
            prompt_path = profile_path.parent / prompt_path
        file_prompt = prompt_path.resolve().read_text(encoding="utf-8").strip()
        system_prompt = (system_prompt + "\n\n" + file_prompt).strip()

    def optional_int(field_name: str) -> int | None:
        value = data.get(field_name)
        if value is None or value == "":
            return None
        return int(value)

    toolsets = validate_profile_toolsets(data.get("toolsets"), profile_path)
    input_access = normalize_profile_input_access(data.get("input_access"), profile_path)
    result_mode = normalize_profile_result_mode(data.get("result_mode"), profile_path)
    self_check_policy = normalize_profile_self_check_policy(
        data.get("self_check_policy"), profile_path
    )
    if input_access != "all" and "file_read" in toolsets:
        raise ValueError(
            "Profiles with input_access other than all must not include the "
            f"file_read toolset: {profile_path}"
        )

    return DeepAgentProfile(
        id=profile_id,
        tool_name=tool_name,
        description=description,
        toolsets=toolsets,
        expose_to_parent=optional_bool(
            data.get("expose_to_parent"), "expose_to_parent", True
        ),
        input_access=input_access,
        system_prompt=system_prompt,
        skill_source_specs=as_string_list(data.get("skill_sources"), "skill_sources"),
        include_global_skills=bool(data.get("include_global_skills", False)),
        image=str(data.get("image") or "").strip(),
        deep_model=str(data.get("deep_model") or "").strip(),
        vision_model=str(data.get("vision_model") or "").strip(),
        deep_recursion_limit=optional_int("deep_recursion_limit"),
        max_review_rounds=optional_int("max_review_rounds"),
        graceful_finalize=normalize_profile_graceful_finalize(
            data.get("graceful_finalize"), profile_path
        ),
        result_mode=result_mode,
        self_check_policy=self_check_policy,
        source_path=str(profile_path),
    )


def load_deep_agent_profiles(
    profile_paths: list[str],
    profile_dirs: list[str],
) -> list[DeepAgentProfile]:
    paths = [
        (Path(path).expanduser().resolve(), True)
        for path in profile_paths
    ]
    for directory in profile_dirs:
        root = Path(directory).expanduser().resolve()
        if not root.exists() or not root.is_dir():
            raise FileNotFoundError(f"Deep Agent profile directory does not exist: {root}")
        for suffix in ("*.json", "*.yaml", "*.yml"):
            paths.extend((path, False) for path in sorted(root.glob(suffix)))

    profiles: list[DeepAgentProfile] = []
    seen_ids: set[str] = set()
    seen_tools: set[str] = set()
    for path, explicit in paths:
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"Deep Agent profile does not exist: {path}")
        profile = load_deep_agent_profile(path)
        if not explicit and not profile.expose_to_parent:
            continue
        if profile.id in seen_ids:
            raise ValueError(f"Duplicate Deep Agent profile id: {profile.id}")
        if profile.tool_name in seen_tools:
            raise ValueError(f"Duplicate Deep Agent tool_name: {profile.tool_name}")
        seen_ids.add(profile.id)
        seen_tools.add(profile.tool_name)
        profiles.append(profile)
    return profiles


def is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def require_safe_output_dir(output_dir: Path, *, sandbox_backend: str = "podman") -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    resolved = output_dir.resolve()
    outputs_root = (
        Path(os.getenv("RUNS_ROOT", "/srv/sandbox-tool/runs")).resolve()
        if sandbox_backend == "controller"
        else OUTPUTS_ROOT.resolve()
    )
    if resolved == outputs_root:
        raise ValueError(f"Refusing to use the run root itself as --output-dir: {outputs_root}")
    if not is_relative_to(resolved, outputs_root):
        raise ValueError(f"--output-dir must be under {outputs_root}: {resolved}")
    return resolved


def prepare_run_directories(run_root: Path) -> dict[str, Path]:
    dirs = {
        "run_root": run_root,
        "input_dir": run_root / "input",
        "workspace_dir": run_root / "workspace",
        "raw_output_dir": run_root / "raw_outputs",
        "clean_export_dir": run_root / "clean_exports",
        "quarantine_dir": run_root / "quarantine",
        "gate_log_dir": run_root / "gate_logs",
        "runner_log_dir": run_root / "runner_logs",
    }
    for path in dirs.values():
        path.mkdir(parents=True, exist_ok=True)
    return {name: path.resolve() for name, path in dirs.items()}


def clear_directory_contents(directory: Path) -> None:
    root = directory.resolve()
    for child in directory.iterdir():
        resolved = child.resolve()
        if not is_relative_to(resolved, root):
            raise RuntimeError(f"Refusing to delete outside output dir: {resolved}")
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def normalize_expected_artifact(path: str) -> str:
    raw = path.replace("\\", "/")
    if not raw.startswith("/outputs/"):
        raise ValueError(f"Expected artifacts must be under /outputs/: {path}")
    normalized = posixpath.normpath(raw)
    if normalized == "/outputs" or normalized.startswith("/outputs/../"):
        raise ValueError(f"Invalid expected artifact path: {path}")
    suffix = posixpath.splitext(normalized)[1].lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError(
            "Expected/review artifacts must use an output-gate allowed extension "
            f"({ALLOWED_EXPORT_EXTENSIONS_TEXT}): {path}"
        )
    return normalized


def normalize_tool_expected_artifacts(
    paths: list[str],
    *,
    allow_empty: bool = False,
) -> list[str]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    normalized = [normalize_expected_artifact(path) for path in paths]
    if not normalized and not allow_empty:
        raise ValueError("expected_artifacts must include at least one artifact")
    if not normalized:
        return []
    allowed_final = set(CONFIG.expected_artifacts)
    invalid = [
        path
        for path in normalized
        if path not in allowed_final and not path.startswith("/outputs/subtasks/")
    ]
    if invalid:
        raise ValueError(
            "Expected artifacts passed to the Deep Agent tool must be either the "
            f"configured final artifacts or intermediate /outputs/subtasks/* artifacts: {invalid}"
        )
    return normalized


def sandbox_input_relative(target: str) -> str:
    raw = target.replace("\\", "/")
    if raw.startswith("/input/"):
        raw = raw.removeprefix("/input/")
    elif raw.startswith("input/"):
        raw = raw.removeprefix("input/")
    raw = posixpath.normpath(raw)
    if raw in {"", "."} or raw.startswith("../") or raw == ".." or raw.startswith("/"):
        raise ValueError(f"Invalid /input target path: {target}")
    return raw


def parse_input_spec(spec: str) -> tuple[Path, str]:
    if "=" in spec:
        host, target = spec.split("=", 1)
    else:
        host = spec
        target = Path(spec).name
    host_path = Path(host).expanduser().resolve()
    if not host_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {host_path}")
    rel = sandbox_input_relative(target)
    return host_path, "/input/" + rel


def stage_inputs(input_specs: list[str], staging_dir: Path) -> list[InputMapping]:
    if staging_dir.exists():
        shutil.rmtree(staging_dir)
    staging_dir.mkdir(parents=True, exist_ok=True)
    mappings: list[InputMapping] = []
    for spec in input_specs:
        host_path, sandbox_path = parse_input_spec(spec)
        rel = sandbox_path.removeprefix("/input/")
        destination = staging_dir / rel
        destination.parent.mkdir(parents=True, exist_ok=True)
        if host_path.is_dir():
            shutil.copytree(host_path, destination)
        else:
            shutil.copy2(host_path, destination)
        mappings.append(
            InputMapping(
                host_path=host_path,
                sandbox_path=sandbox_path,
                staged_path=destination,
            )
        )
    return mappings


def stage_skill_sources(
    skill_specs: list[str],
    input_dir: Path,
    *,
    base_dir: Path | None = None,
) -> list[str]:
    """Stage host skill source directories into /input and return sandbox source paths."""
    staged_sources: list[str] = []
    for spec in skill_specs:
        if "=" in spec:
            host, target = spec.split("=", 1)
        else:
            host = spec
            target = ""
        host_path = Path(host).expanduser()
        if not host_path.is_absolute() and base_dir is not None:
            host_path = base_dir / host_path
        host_path = host_path.resolve()
        if not host_path.exists() or not host_path.is_dir():
            raise FileNotFoundError(f"Skill source must be an existing directory: {host_path}")
        is_single_skill = (host_path / "SKILL.md").is_file()
        if not target:
            target = f"skills/{host_path.name}" if is_single_skill else "skills"
        rel = sandbox_input_relative(target)
        destination = input_dir / rel
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.exists():
            shutil.rmtree(destination)
        shutil.copytree(host_path, destination)
        source_rel = posixpath.dirname(rel) if is_single_skill else rel
        if source_rel in {"", "."}:
            raise ValueError(
                "Single skill sources must be staged below a source directory, "
                f"for example {host_path}=skills/{host_path.name}"
            )
        staged_sources.append("/input/" + source_rel)
    return list(dict.fromkeys(staged_sources))


def stage_profile_skill_sources(profile: DeepAgentProfile, input_dir: Path) -> list[str]:
    staged_sources: list[str] = []
    host_specs: list[str] = []
    for spec in profile.skill_source_specs:
        normalized = spec.replace("\\", "/")
        if normalized == "/input" or normalized.startswith("/input/"):
            staged_sources.append(posixpath.normpath(normalized))
        else:
            host_specs.append(spec)
    base_dir = Path(profile.source_path).parent if profile.source_path else None
    staged_sources.extend(stage_skill_sources(host_specs, input_dir, base_dir=base_dir))
    return staged_sources


def materialize_deep_agent_profiles(
    profiles: list[DeepAgentProfile],
    input_dir: Path,
    global_skill_sources: list[str],
) -> None:
    for profile in profiles:
        profile_sources = stage_profile_skill_sources(profile, input_dir)
        combined_sources = [
            *(global_skill_sources if profile.include_global_skills else []),
            *profile_sources,
        ]
        profile.skill_sources = list(dict.fromkeys(combined_sources))


def resolve_sandbox_path(path: str | Path) -> Path:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    raw = str(path).replace("\\", "/")
    if raw == "/outputs":
        return CONFIG.output_dir
    if raw.startswith("/outputs/"):
        return CONFIG.output_dir / raw.removeprefix("/outputs/")
    if raw == "/exports":
        return CONFIG.clean_export_dir
    if raw.startswith("/exports/"):
        return CONFIG.clean_export_dir / raw.removeprefix("/exports/")
    if raw == "/input":
        return CONFIG.input_dir
    if raw.startswith("/input/"):
        return CONFIG.input_dir / raw.removeprefix("/input/")
    return Path(path)


def normalize_readable_virtual_path(path: str | Path) -> str:
    raw = str(path).replace("\\", "/")
    normalized = posixpath.normpath(raw)
    if normalized in {"/outputs", "/exports", "/input"}:
        return normalized
    if (
        normalized.startswith("/outputs/")
        or normalized.startswith("/exports/")
        or normalized.startswith("/input/")
    ):
        return normalized
    raise ValueError(f"Path must be under /input, /outputs, or /exports: {path}")


def resolve_readable_virtual_path(path: str | Path) -> tuple[str, Path]:
    normalized = normalize_readable_virtual_path(path)
    host_path = resolve_sandbox_path(normalized).resolve()
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    allowed_roots = [
        CONFIG.output_dir.resolve(),
        CONFIG.clean_export_dir.resolve(),
        CONFIG.input_dir.resolve(),
    ]
    if not any(host_path == root or is_relative_to(host_path, root) for root in allowed_roots):
        raise ValueError(f"Resolved path escaped /input or /outputs: {path}")
    return normalized, host_path


def timestamp_iso(path: Path) -> str | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")


def artifact_info(path: str | Path) -> dict[str, Any]:
    host_path = resolve_sandbox_path(path)
    info: dict[str, Any] = {
        "sandbox_path": str(path),
        "host_path": str(host_path),
        "exists": host_path.exists(),
    }
    if host_path.exists():
        info["mtime"] = timestamp_iso(host_path)
        info["is_file"] = host_path.is_file()
        info["is_dir"] = host_path.is_dir()
        if host_path.is_file():
            info["bytes"] = host_path.stat().st_size
    return info


def check_expected_artifacts(paths: list[str | Path]) -> dict[str, Any]:
    artifacts = [artifact_info(path) for path in paths]
    return {
        "artifacts": artifacts,
        "ok": all(item["exists"] for item in artifacts),
        "missing": [item for item in artifacts if not item["exists"]],
    }


def review_markdown_blockers(text: str, artifact_name: str) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    active_patterns = [
        (r"<\s*/?\s*script\b", "literal_script_tag"),
        (r"<\s*iframe\b", "literal_iframe_tag"),
        (r"<\s*object\b", "literal_object_tag"),
        (r"<\s*embed\b", "literal_embed_tag"),
        (r"javascript\s*:", "javascript_url"),
    ]
    for pattern, code in active_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            line = text[: match.start()].count("\n") + 1
            findings.append(
                {
                    "artifact": artifact_name,
                    "code": code,
                    "line": line,
                    "message": (
                        "Markdown review artifacts must escape active-content "
                        "examples, e.g. use &lt;script&gt; or plain 'script tag'."
                    ),
                }
            )
    return findings


def validate_review_artifact_preflight(artifacts: list[str]) -> list[dict[str, Any]]:
    findings: list[dict[str, Any]] = []
    for artifact in artifacts:
        try:
            path = resolve_sandbox_path(artifact)
        except Exception as exc:
            findings.append(
                {
                    "artifact": artifact,
                    "code": "invalid_path",
                    "message": str(exc),
                }
            )
            continue
        if not path.is_file():
            findings.append(
                {
                    "artifact": artifact,
                    "code": "missing_review_artifact",
                    "message": "Review artifact does not exist or is not a file.",
                }
            )
            continue
        if path.stat().st_size == 0:
            findings.append(
                {
                    "artifact": artifact,
                    "code": "empty_review_artifact",
                    "message": "Review artifact is empty.",
                }
            )
            continue
        if path.suffix.lower() == ".md":
            text = path.read_text(encoding="utf-8", errors="replace")
            findings.extend(review_markdown_blockers(text, path.name))
            if path.name == "self_check_report.md":
                lower = text.lower()
                if any(
                    marker in lower
                    for marker in (
                        "pending",
                        "not yet executed",
                        "not executed",
                        "未実行",
                    )
                ):
                    findings.append(
                        {
                            "artifact": artifact,
                            "code": "self_check_report_pending",
                            "message": (
                                "Self-check report appears pending or not executed. "
                                "Run the self-check and update the report before review."
                            ),
                        }
                    )
    return findings


def build_unreviewed_artifact_salvage(
    *,
    profile: DeepAgentProfile,
    effective_expected_artifacts: list[str],
    artifact_check: dict[str, Any],
    deep_error: dict[str, Any] | None,
    review_requested: bool,
    self_check_dir: Path,
) -> dict[str, Any]:
    if profile.result_mode == "inline" or deep_error is None or review_requested:
        return {"available": False}
    existing_expected = [
        item.get("sandbox_path")
        for item in artifact_check.get("artifacts", [])
        if item.get("exists") and item.get("is_file")
    ]
    self_check_candidates: list[str] = []
    for name in ("self_check_plan.md", "self_check_report.md"):
        if (self_check_dir / name).exists():
            self_check_candidates.append(f"/outputs/subtasks/{name}")
    candidate_review_artifacts = list(
        dict.fromkeys(
            [
                *(path for path in effective_expected_artifacts if path in existing_expected),
                *self_check_candidates,
            ]
        )
    )
    if not candidate_review_artifacts:
        return {"available": False}
    return {
        "available": True,
        "reason": "deep_agent_error_before_parent_review",
        "deep_error": deep_error,
        "existing_expected_artifacts": existing_expected,
        "self_check_artifacts": self_check_candidates,
        "candidate_review_artifacts": candidate_review_artifacts,
        "review_requested": False,
        "gate_allowed": False,
        "message": (
            "Artifacts exist after a Deep Agent error, but request_parent_review "
            "was not called. Treat them as unreviewed candidates only. Do not run "
            "output gate until a later Deep Agent attempt explicitly requests review."
        ),
    }


def wsl_podman_image_available(image: str, distro: str) -> bool:
    if shutil.which("wsl") is None:
        return False
    return (
        subprocess.run(
            ["wsl", "-d", distro, "--", "sudo", "podman", "image", "exists", image],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        ).returncode
        == 0
    )


def native_podman_image_available(image: str, podman_bin: str) -> bool:
    if shutil.which(podman_bin) is None:
        return False
    return (
        subprocess.run(
            [podman_bin, "image", "exists", image],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        ).returncode
        == 0
    )


def resolve_host_os(host_os: str) -> str:
    if host_os != "auto":
        return host_os
    return "windows" if os.name == "nt" else "linux"


def configured_image_available(image: str | None = None) -> bool:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    image_name = image or CONFIG.image
    if CONFIG.sandbox_backend == "controller":
        return True
    if CONFIG.host_os == "windows":
        return wsl_podman_image_available(image_name, CONFIG.wsl_distro)
    return native_podman_image_available(image_name, CONFIG.podman_bin)


def profile_input_dir(profile: DeepAgentProfile) -> Path:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    if profile.input_access == "all":
        return CONFIG.input_dir

    root = CONFIG.run_root / "profile_inputs" / safe_tool_name(profile.id)
    if root.exists():
        shutil.rmtree(root)
    root.mkdir(parents=True, exist_ok=True)
    if profile.input_access == "skills_only":
        for virtual_source in profile.skill_sources:
            normalized = virtual_source.replace("\\", "/").strip().lstrip("/")
            if not normalized.startswith("input/"):
                continue
            rel = normalized.removeprefix("input/").strip("/")
            if not rel:
                continue
            source = (CONFIG.input_dir / rel).resolve()
            destination = (root / rel).resolve()
            if not is_relative_to(source, CONFIG.input_dir.resolve()):
                raise ValueError(f"Profile skill source escapes input dir: {virtual_source}")
            if source.is_dir():
                shutil.copytree(source, destination, dirs_exist_ok=True)
            elif source.is_file():
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, destination)
    return root


def create_configured_backend(
    image: str | None = None,
    *,
    input_dir: Path | None = None,
) -> Any:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    image_name = image or CONFIG.image
    effective_input_dir = input_dir or CONFIG.input_dir

    if CONFIG.sandbox_backend == "controller":
        if not CONFIG.sandbox_controller_url:
            raise RuntimeError("sandbox_controller_url is required for controller backend")
        CONFIG.workspace_dir.mkdir(parents=True, exist_ok=True)
        return ControllerSandboxBackend(
            image=image_name,
            controller_url=CONFIG.sandbox_controller_url,
            run_id=CONFIG.run_root.name,
            input_dir=effective_input_dir,
            output_dir=CONFIG.output_dir,
            workspace_dir=CONFIG.workspace_dir,
            token=CONFIG.sandbox_controller_token,
            timeout_seconds=300,
            max_output_bytes=450_000,
        )

    limits = PodmanRunLimits(
        enforce_cgroups=CONFIG.host_os != "windows",
        shell_cpu_seconds=300,
        shell_virtual_memory_kb=4 * 1024 * 1024,
    )

    if CONFIG.host_os == "windows":
        return PodmanSandboxBackend.for_wsl(
            image=image_name,
            distro=CONFIG.wsl_distro,
            input_dir=effective_input_dir,
            output_dir=CONFIG.output_dir,
            use_sudo=CONFIG.wsl_use_sudo,
            limits=limits,
            max_output_bytes=450_000,
        )

    from podman_sandbox_backend import PodmanSecurityOptions

    return PodmanSandboxBackend(
        image=image_name,
        input_dir=effective_input_dir,
        output_dir=CONFIG.output_dir,
        podman=CONFIG.podman_bin,
        host_path_mode="native",
        limits=limits,
        security=PodmanSecurityOptions(
            pull="never",
            selinux_relabel=CONFIG.selinux_relabel,
            userns="keep-id",
            user=None,
        ),
        max_output_bytes=450_000,
    )


def message_role(message: Any) -> str:
    return getattr(message, "type", message.__class__.__name__)


def content_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    return json.dumps(content, ensure_ascii=False, default=str)


def tool_call_dicts(message: Any) -> list[dict[str, Any]]:
    tool_calls = getattr(message, "tool_calls", None) or []
    normalized = []
    for call in tool_calls:
        if isinstance(call, dict):
            normalized.append({"name": call.get("name", ""), "args": call.get("args", {})})
        else:
            normalized.append(
                {"name": getattr(call, "name", ""), "args": getattr(call, "args", {})}
            )
    return normalized


def trace_messages(messages: list[Any], *, content_limit: int = 2200) -> list[dict[str, Any]]:
    trace = []
    for index, message in enumerate(messages, start=1):
        calls = tool_call_dicts(message)
        trace.append(
            {
                "index": index,
                "role": message_role(message),
                "tool_calls": [call["name"] for call in calls],
                "tool_call_args": calls,
                "content_preview": content_text(getattr(message, "content", ""))[:content_limit],
            }
        )
    return trace


def _skill_records_from_text(text: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for match in SKILL_VIRTUAL_PATH_RE.finditer(text):
        path = match.group("path")
        skill = match.group("skill")
        key = (skill, path)
        if key in seen:
            continue
        seen.add(key)
        records.append({"skill": skill, "path": path})
    return records


def extract_skill_usage_from_trace(
    trace: list[dict[str, Any]],
    configured_sources: list[str] | None = None,
) -> dict[str, Any]:
    """Summarize whether a Deep Agent actually referenced or executed skill files."""
    references: list[dict[str, Any]] = []
    executions: list[dict[str, Any]] = []
    skill_names: set[str] = set()
    seen_references: set[tuple[Any, str, str, str]] = set()
    seen_executions: set[tuple[Any, str, str, str]] = set()

    for item in trace:
        trace_index = item.get("index")
        for call in item.get("tool_call_args", []):
            tool_name = str(call.get("name") or "")
            args = call.get("args") if isinstance(call, dict) else {}
            if isinstance(args, str):
                scan_text = args
                command = args if tool_name == "execute" else ""
            else:
                scan_text = json.dumps(args, ensure_ascii=False, default=str)
                command = str(args.get("command") or "") if isinstance(args, dict) else ""

            for record in _skill_records_from_text(scan_text):
                skill_names.add(record["skill"])
                base = {
                    "trace_index": trace_index,
                    "tool": tool_name,
                    "skill": record["skill"],
                    "path": record["path"],
                }
                if tool_name == "execute":
                    execution = {**base, "command": command[:1000]}
                    key = (
                        trace_index,
                        execution["tool"],
                        execution["skill"],
                        execution["path"],
                    )
                    if key not in seen_executions:
                        seen_executions.add(key)
                        executions.append(execution)
                else:
                    key = (trace_index, base["tool"], base["skill"], base["path"])
                    if key not in seen_references:
                        seen_references.add(key)
                        references.append(base)

    return {
        "configured": bool(configured_sources),
        "configured_sources": configured_sources or [],
        "referenced": bool(references),
        "executed": bool(executions),
        "skill_names": sorted(skill_names),
        "references": references[:50],
        "executions": executions[:50],
    }


def input_manifest() -> list[dict[str, str]]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    return [
        {
            "host_path": str(item.host_path),
            "sandbox_path": item.sandbox_path,
            "staged_path": str(item.staged_path),
        }
        for item in CONFIG.input_mappings
    ]


def tool_names_for_toolsets(toolsets: list[str]) -> list[str]:
    names: list[str] = []
    for toolset in toolsets:
        for tool_name in TOOLSET_TOOL_NAMES[toolset]:
            if tool_name not in names:
                names.append(tool_name)
    return names


def tool_names_for_profile(profile: DeepAgentProfile) -> list[str]:
    names = tool_names_for_toolsets(profile.toolsets)
    if profile.result_mode == "inline":
        names = [name for name in names if name != "request_parent_review"]
    return names


def profile_summary_for_prompt(profile: DeepAgentProfile) -> str:
    model_note = f"; model={profile.deep_model}" if profile.deep_model else ""
    vision_model_note = (
        f"; vision_model={profile.vision_model}" if profile.vision_model else ""
    )
    image_note = f"; image={profile.image}" if profile.image else ""
    toolset_note = f"; toolsets={', '.join(profile.toolsets)}" if profile.toolsets else ""
    tools_note = (
        f"; custom_tools={', '.join(tool_names_for_profile(profile))}"
        if profile.toolsets
        else ""
    )
    rounds_note = (
        f"; max_review_rounds={profile.max_review_rounds}"
        if profile.max_review_rounds is not None
        else ""
    )
    input_note = f"; input_access={profile.input_access}"
    result_note = f"; result_mode={profile.result_mode}"
    self_check_note = f"; self_check_policy={profile.self_check_policy}"
    graceful_note = (
        f"; graceful_finalize={json.dumps(profile.graceful_finalize, ensure_ascii=False)}"
        if profile.graceful_finalize
        else ""
    )
    skill_note = (
        f"; skills={', '.join(profile.skill_sources)}" if profile.skill_sources else ""
    )
    return (
        f"- {profile.tool_name}: {profile.description} "
        f"(profile_id={profile.id}{toolset_note}{tools_note}{model_note}"
        f"{vision_model_note}{image_note}{rounds_note}{input_note}{result_note}"
        f"{self_check_note}{graceful_note}{skill_note})"
    )


def deep_agent_profile_prompt_section() -> str:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    if CONFIG.deep_agent_profiles:
        return (
            "\n\nAvailable Deep Agent profile tools:\n"
            + "\n".join(
                profile_summary_for_prompt(profile)
                for profile in CONFIG.deep_agent_profiles
            )
            + "\nChoose the most appropriate profile tool for the requested work. "
            "Use the same or a better-suited profile for repair attempts. "
            "If a profile has input_access=skills_only or input_access=none, "
            "do not send it tasks that require reading user-provided /input files; "
            "first use a non-network profile to extract the needed non-sensitive "
            "values into an intermediate artifact."
        )
    if CONFIG.skill_sources:
        return (
            "\n\nDeep Agent skill sources:\n"
            + "\n".join(f"- {path}" for path in CONFIG.skill_sources)
        )
    return ""


def build_parent_prompt() -> str:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    delegate_text = (
        "the most appropriate Deep Agent profile tool"
        if CONFIG.deep_agent_profiles
        else "the Deep Agent tool"
    )
    attempt_limit_text = (
        f"Profile tools may define their own max_review_rounds; runner default is {CONFIG.max_review_rounds}."
        if CONFIG.deep_agent_profiles
        else f"Maximum Deep Agent attempts allowed: {CONFIG.max_review_rounds}"
    )
    expected_text = "\n".join(f"- {path}" for path in CONFIG.expected_artifacts)
    if not expected_text:
        expected_text = "(none configured; inline-result profiles may be used)"
    return (
        f"Run this task by delegating to {delegate_text}. Artifact-result profiles must "
        "request parent review through request_parent_review before the parent can close "
        "the task. Inline-result profiles do not create final artifacts, do not request "
        "parent review, and should be closed from their inline_result when the tool "
        "returns ok=true.\n\n"
        "Inputs available in the sandbox:\n"
        + "\n".join(f"- {item.sandbox_path}" for item in CONFIG.input_mappings)
        + "\n\nExpected artifacts:\n"
        + expected_text
        + "\n\nOutput-gate allowed review/export extensions:\n"
        + ALLOWED_EXPORT_EXTENSIONS_TEXT
        + "\nDo not ask the Deep Agent to create final or review artifacts with any "
        "other extension. Helper scripts or working files may exist under /outputs, "
        "but they must not be passed as expected_artifacts or request_parent_review artifacts. "
        "When you call an artifact-result Deep Agent profile tool, pass the configured "
        "final artifacts exactly as expected_artifacts. When you call an inline-result "
        "profile tool, pass expected_artifacts=[] and do not run output gate afterward. "
        "If you ask an artifact profile for additional self-check plan/report review "
        "artifacts, place them under /outputs/subtasks/ and include them in the Deep "
        "Agent task text, not as root-level expected_artifacts."
        + deep_agent_profile_prompt_section()
        + f"\n\n{attempt_limit_text}"
        + "\n\nUser task:\n"
        + CONFIG.prompt.strip()
    )


def text_file_preview(path: Path, max_chars: int) -> dict[str, Any]:
    size = path.stat().st_size
    max_bytes = max(max_chars * 4, 4096)
    with path.open("rb") as handle:
        raw = handle.read(max_bytes + 1)
    truncated = len(raw) > max_bytes or size > max_bytes
    text = raw[:max_bytes].decode("utf-8", errors="replace")
    if len(text) > max_chars:
        text = text[:max_chars]
        truncated = True
    return {
        "kind": "text",
        "bytes": size,
        "truncated": truncated,
        "preview": text,
    }


def csv_file_preview(path: Path, max_rows: int, max_cols: int) -> dict[str, Any]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            if index >= max_rows:
                break
            rows.append(row[:max_cols])
    return {
        "kind": "csv",
        "bytes": path.stat().st_size,
        "delimiter": delimiter,
        "preview_rows": rows,
        "preview_row_count": len(rows),
        "max_cols_returned": max_cols,
    }


def json_file_preview(path: Path, max_chars: int) -> dict[str, Any]:
    preview = text_file_preview(path, max_chars=max_chars)
    try:
        parsed = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        preview["json_parse_error"] = f"{exc.__class__.__name__}: {str(exc)[:500]}"
        return preview

    if isinstance(parsed, dict):
        preview["json_type"] = "object"
        preview["top_level_keys"] = list(parsed.keys())[:50]
    elif isinstance(parsed, list):
        preview["json_type"] = "array"
        preview["array_length"] = len(parsed)
        preview["first_item_type"] = type(parsed[0]).__name__ if parsed else None
    else:
        preview["json_type"] = type(parsed).__name__
    return preview


def image_file_preview(path: Path) -> dict[str, Any]:
    mime = image_mime_type(path)
    info: dict[str, Any] = {
        "kind": "image",
        "bytes": path.stat().st_size,
        "mime": mime,
        "semantic_read_supported": mime.startswith("image/"),
        "semantic_read_tool": "read_sandbox_file(question=...) or inspect_sandbox_image",
    }
    try:
        from PIL import Image
    except Exception as exc:  # pragma: no cover - depends on local environment
        info["metadata_error"] = f"pillow_unavailable: {exc.__class__.__name__}: {exc}"
        return info

    try:
        with Image.open(path) as image:
            info.update(
                {
                    "format": image.format,
                    "width": image.width,
                    "height": image.height,
                    "mode": image.mode,
                    "frames": getattr(image, "n_frames", 1),
                    "has_alpha": image.mode in {"RGBA", "LA"}
                    or ("transparency" in image.info),
                }
            )
    except Exception as exc:
        info["metadata_error"] = f"{exc.__class__.__name__}: {exc}"
    return info


def pdf_file_preview(path: Path, max_chars: int, max_pages: int) -> dict[str, Any]:
    info: dict[str, Any] = {
        "kind": "pdf",
        "bytes": path.stat().st_size,
        "text_extraction": "pypdf",
    }
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover - depends on local environment
        info["error"] = f"pypdf_unavailable: {exc.__class__.__name__}: {exc}"
        return info

    try:
        reader = PdfReader(str(path))
        info["page_count"] = len(reader.pages)
        if reader.metadata:
            info["metadata"] = {
                str(key).lstrip("/"): str(value)
                for key, value in reader.metadata.items()
                if value is not None
            }

        remaining_chars = max_chars
        page_previews: list[dict[str, Any]] = []
        pages_to_scan = min(max_pages, len(reader.pages))
        for page_index in range(pages_to_scan):
            if remaining_chars <= 0:
                break
            try:
                text = reader.pages[page_index].extract_text() or ""
            except Exception as exc:
                page_previews.append(
                    {
                        "page": page_index + 1,
                        "error": f"{exc.__class__.__name__}: {exc}",
                    }
                )
                continue
            if len(text) > remaining_chars:
                text = text[:remaining_chars]
            remaining_chars -= len(text)
            page_previews.append(
                {
                    "page": page_index + 1,
                    "chars_returned": len(text),
                    "text_preview": text,
                }
            )

        info.update(
            {
                "pages_scanned": pages_to_scan,
                "preview_pages": page_previews,
                "truncated": len(reader.pages) > pages_to_scan or remaining_chars <= 0,
            }
        )
        if not any(page.get("text_preview") for page in page_previews):
            info["note"] = (
                "No extractable text was found in the scanned preview pages. "
                "For scanned PDFs, render pages to images inside the sandbox and "
                "call read_sandbox_file with a visual question on those images."
            )
    except Exception as exc:
        info["error"] = f"{exc.__class__.__name__}: {exc}"
    return info


def media_file_preview(path: Path) -> dict[str, Any]:
    mime, _ = mimetypes.guess_type(path.name)
    kind = "binary_or_unsupported"
    if mime:
        if mime.startswith("audio/"):
            kind = "audio"
        elif mime.startswith("video/"):
            kind = "video"
        elif mime.startswith("application/"):
            kind = "application_binary"
    return {
        "kind": kind,
        "bytes": path.stat().st_size,
        "mime": mime or "application/octet-stream",
        "semantic_read_supported": False,
        "note": (
            "This runner reports metadata for this file type. If semantic reading is "
            "needed, create a task-specific extractor inside the sandbox and save a "
            "text/JSON artifact for review."
        ),
    }


def xlsx_file_preview(path: Path, max_rows: int, max_cols: int) -> dict[str, Any]:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - depends on local environment
        return {
            "kind": "xlsx",
            "bytes": path.stat().st_size,
            "error": f"openpyxl_unavailable: {exc.__class__.__name__}: {exc}",
        }

    workbook_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    workbook_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
    sheets: list[dict[str, Any]] = []
    error_strings: list[dict[str, str]] = []
    scan_cell_limit = 200_000

    try:
        for sheet_name in workbook_values.sheetnames:
            ws_values = workbook_values[sheet_name]
            ws_formulas = workbook_formulas[sheet_name]
            preview_rows: list[list[Any]] = []
            nonempty_cells = 0
            scanned_cells = 0
            scan_truncated = False

            for row_index, row in enumerate(ws_values.iter_rows(values_only=True), start=1):
                row_values = list(row[:max_cols])
                if row_index <= max_rows:
                    preview_rows.append(row_values)
                for value in row:
                    scanned_cells += 1
                    if value not in (None, ""):
                        nonempty_cells += 1
                    if scanned_cells >= scan_cell_limit:
                        scan_truncated = True
                        break
                if scan_truncated:
                    break

            error_scan_count = 0
            for row in ws_formulas.iter_rows():
                for cell in row:
                    error_scan_count += 1
                    value = cell.value
                    if isinstance(value, str) and value.startswith("#"):
                        error_strings.append(
                            {"sheet": sheet_name, "cell": cell.coordinate, "value": value}
                        )
                        if len(error_strings) >= 100:
                            break
                    if error_scan_count >= scan_cell_limit:
                        break
                if len(error_strings) >= 100 or error_scan_count >= scan_cell_limit:
                    break

            sheets.append(
                {
                    "name": sheet_name,
                    "max_row": ws_values.max_row,
                    "max_column": ws_values.max_column,
                    "nonempty_cells_scanned": nonempty_cells,
                    "scan_truncated": scan_truncated,
                    "preview_rows": preview_rows,
                }
            )
    finally:
        workbook_values.close()
        workbook_formulas.close()

    return {
        "kind": "xlsx",
        "bytes": path.stat().st_size,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "excel_error_strings": error_strings,
        "excel_error_count_returned": len(error_strings),
    }


def inspect_file_by_type(
    path: Path,
    max_chars: int,
    max_rows: int,
    max_cols: int,
    max_pages: int = 5,
) -> dict[str, Any]:
    if not path.exists():
        return {"exists": False}
    info: dict[str, Any] = {
        "exists": True,
        "is_file": path.is_file(),
        "is_dir": path.is_dir(),
        "mtime": timestamp_iso(path),
    }
    if path.is_dir():
        children = sorted(path.iterdir(), key=lambda child: child.name.lower())[:100]
        info["children"] = [
            {
                "name": child.name,
                "is_file": child.is_file(),
                "is_dir": child.is_dir(),
                "bytes": child.stat().st_size if child.is_file() else None,
            }
            for child in children
        ]
        info["child_count_returned"] = len(children)
        return info

    suffix = path.suffix.lower()
    if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tif", ".tiff"}:
        info.update(image_file_preview(path))
    elif suffix == ".pdf":
        info.update(pdf_file_preview(path, max_chars=max_chars, max_pages=max_pages))
    elif suffix in {".xlsx", ".xlsm"}:
        info.update(xlsx_file_preview(path, max_rows=max_rows, max_cols=max_cols))
    elif suffix in {".csv", ".tsv"}:
        info.update(csv_file_preview(path, max_rows=max_rows, max_cols=max_cols))
    elif suffix == ".json":
        info.update(json_file_preview(path, max_chars=max_chars))
    elif suffix in {
        ".txt",
        ".md",
        ".log",
        ".py",
        ".js",
        ".ts",
        ".html",
        ".css",
        ".xml",
        ".yaml",
        ".yml",
    }:
        info.update(text_file_preview(path, max_chars=max_chars))
    else:
        info.update(media_file_preview(path))
    return info


@tool
def list_sandbox_files(root: str = "/outputs", max_entries: int = 100) -> dict[str, Any]:
    """List files under /input or /outputs for parent review."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        normalized, host_root = resolve_readable_virtual_path(root)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    max_entries = max(1, min(max_entries, 500))
    if not host_root.exists():
        return {"ok": False, "root": normalized, "exists": False, "files": []}

    files: list[dict[str, Any]] = []
    for child in sorted(host_root.rglob("*"), key=lambda item: str(item).lower()):
        if len(files) >= max_entries:
            break
        resolved = child.resolve()
        if not is_relative_to(resolved, host_root.resolve()):
            continue
        rel = child.relative_to(host_root).as_posix()
        virtual_path = normalized.rstrip("/") + "/" + rel
        files.append(
            {
                "path": virtual_path,
                "is_file": child.is_file(),
                "is_dir": child.is_dir(),
                "bytes": child.stat().st_size if child.is_file() else None,
                "mtime": timestamp_iso(child),
            }
        )
    return {
        "ok": True,
        "root": normalized,
        "host_root": str(host_root),
        "entries_returned": len(files),
        "truncated": len(files) >= max_entries,
        "files": files,
    }


@tool
def inspect_sandbox_file(
    path: str,
    max_chars: int = 12000,
    max_rows: int = 20,
    max_cols: int = 12,
    max_pages: int = 5,
) -> dict[str, Any]:
    """Inspect a file under /input or /outputs, including text/csv/json/xlsx/pdf/image previews."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        normalized, host_path = resolve_readable_virtual_path(path)
        max_chars = max(1000, min(max_chars, 50000))
        max_rows = max(1, min(max_rows, 100))
        max_cols = max(1, min(max_cols, 50))
        max_pages = max(1, min(max_pages, 20))
        info = inspect_file_by_type(
            host_path,
            max_chars=max_chars,
            max_rows=max_rows,
            max_cols=max_cols,
            max_pages=max_pages,
        )
        info.update(
            {
                "ok": bool(info.get("exists", True)),
                "path": normalized,
                "host_path": str(host_path),
            }
        )
        return info
    except Exception as exc:
        return {"ok": False, "path": path, "error": f"{exc.__class__.__name__}: {exc}"}


def openai_model_name(model: str) -> str:
    if model.startswith("openai:"):
        return model.split(":", 1)[1]
    return model


def image_mime_type(path: Path) -> str:
    guessed, _ = mimetypes.guess_type(path.name)
    if guessed and guessed.startswith("image/"):
        return guessed
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".png":
        return "image/png"
    if suffix == ".webp":
        return "image/webp"
    return "application/octet-stream"


def active_vision_model_name() -> str:
    if CONFIG is None:
        return ""
    configured = ACTIVE_DEEP_AGENT_VISION_MODEL or CONFIG.vision_model or CONFIG.deep_model
    return openai_model_name(configured)


def build_tiled_vision_image(
    source: "Image.Image",
    source_digest: str,
    normalized_path: str,
    grid: int,
) -> tuple[bytes, dict[str, Any]]:
    from PIL import Image

    width, height = source.size
    source_rgb = source.convert("RGB")
    tiled = Image.new("RGB", (width * grid, height * grid), "white")
    for y in range(grid):
        for x in range(grid):
            tiled.paste(source_rgb, (x * width, y * height))

    buffer = BytesIO()
    tiled.save(buffer, format="PNG")
    payload = buffer.getvalue()

    digest = hashlib.sha256((normalized_path + source_digest).encode("utf-8")).hexdigest()
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(normalized_path).stem).strip("._") or "image"
    tile_name = f"{safe_stem}_tile{grid}x{grid}_{digest[:12]}.png"
    virtual_path = ""
    if CONFIG is not None and getattr(CONFIG, "output_dir", None):
        tile_dir = CONFIG.output_dir / "_vision_tiles"
        tile_dir.mkdir(parents=True, exist_ok=True)
        tile_path = tile_dir / tile_name
        tile_path.write_bytes(payload)
        virtual_path = f"/outputs/_vision_tiles/{tile_name}"

    return payload, {
        "applied": True,
        "transform": f"tile_{grid}x{grid}",
        "sent_width": width * grid,
        "sent_height": height * grid,
        "virtual_path": virtual_path,
    }


def content_bbox_for_padding(image: "Image.Image", background_threshold: int = 245) -> tuple[int, int, int, int]:
    rgba = image.convert("RGBA")
    arr = np.asarray(rgba)
    alpha = arr[:, :, 3]
    rgb = arr[:, :, :3].astype(np.int16)
    non_white = np.any(rgb < background_threshold, axis=2)
    content = (alpha > 0) & non_white
    ys, xs = np.where(content)
    if len(xs) == 0 or len(ys) == 0:
        return 0, 0, image.width, image.height
    return int(xs.min()), int(ys.min()), int(xs.max()) + 1, int(ys.max()) + 1


def add_padding_if_needed(
    image: "Image.Image",
    min_margin: int = VISION_MIN_MARGIN_PX,
) -> tuple["Image.Image", dict[str, Any]]:
    from PIL import Image

    source = image.convert("RGBA")
    width, height = source.size
    left, top, right, bottom = content_bbox_for_padding(source)
    margins = {
        "left": left,
        "top": top,
        "right": width - right,
        "bottom": height - bottom,
    }
    padding = {
        side: max(0, min_margin - margin)
        for side, margin in margins.items()
    }
    applied = any(value > 0 for value in padding.values())
    if not applied:
        return source.convert("RGB"), {
            "applied": False,
            "min_margin": min_margin,
            "content_bbox": [left, top, right, bottom],
            "margins": margins,
        }

    padded_rgba = Image.new(
        "RGBA",
        (width + padding["left"] + padding["right"], height + padding["top"] + padding["bottom"]),
        (255, 255, 255, 255),
    )
    padded_rgba.paste(source, (padding["left"], padding["top"]), source)
    padded = padded_rgba.convert("RGB")
    return padded, {
        "applied": True,
        "min_margin": min_margin,
        "content_bbox": [left, top, right, bottom],
        "margins": margins,
        "padding": padding,
        "padded_width": padded.width,
        "padded_height": padded.height,
    }


def save_prepared_vision_image(
    image: "Image.Image",
    normalized_path: str,
    source_digest: str,
    suffix: str,
) -> str:
    if CONFIG is None or not getattr(CONFIG, "output_dir", None):
        return ""
    digest = hashlib.sha256((normalized_path + source_digest + suffix).encode("utf-8")).hexdigest()
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(normalized_path).stem).strip("._") or "image"
    file_name = f"{safe_stem}_{suffix}_{digest[:12]}.png"
    prepared_dir = CONFIG.output_dir / "_vision_prepared"
    prepared_dir.mkdir(parents=True, exist_ok=True)
    path = prepared_dir / file_name
    image.convert("RGB").save(path)
    return f"/outputs/_vision_prepared/{file_name}"


def prepare_vision_payload_bytes(
    host_path: Path,
    normalized_path: str,
    mime: str,
) -> tuple[bytes, str, dict[str, Any]]:
    try:
        from PIL import Image

        with Image.open(host_path) as image:
            original = image.convert("RGBA")
        source_digest = sha256_file(host_path)
        original_width, original_height = original.size
        padded, padding_info = add_padding_if_needed(original)
        transform: dict[str, Any] = {
            "applied": bool(padding_info.get("applied")),
            "source_width": original_width,
            "source_height": original_height,
            "padding": padding_info,
        }

        if CONFIG is not None and getattr(CONFIG, "tile_small_images_for_vision", False):
            max_side = max(1, int(getattr(CONFIG, "vision_tile_max_side", 128)))
            grid = max(2, min(int(getattr(CONFIG, "vision_tile_grid", 5)), 10))
            if max(original_width, original_height) <= max_side:
                tiled_bytes, tile_info = build_tiled_vision_image(
                    padded,
                    source_digest,
                    normalized_path,
                    grid,
                )
                transform.update(tile_info)
                transform["applied"] = True
                transform["max_side"] = max_side
                return tiled_bytes, "image/png", transform
            transform["reason"] = "image_not_small"
            transform["max_side"] = max_side

        if padding_info.get("applied"):
            virtual_path = save_prepared_vision_image(
                padded,
                normalized_path,
                source_digest,
                "padded",
            )
            transform["transform"] = "pad_to_min_margin"
            transform["sent_width"] = padded.width
            transform["sent_height"] = padded.height
            transform["virtual_path"] = virtual_path
            buffer = BytesIO()
            padded.save(buffer, format="PNG")
            return buffer.getvalue(), "image/png", transform

        original_bytes = host_path.read_bytes()
        return original_bytes, mime, transform
    except Exception as exc:
        original_bytes = host_path.read_bytes()
        return original_bytes, mime, {"applied": False, "error": f"{exc.__class__.__name__}: {exc}"}


@tool
def inspect_sandbox_image(
    path: str,
    question: str,
    max_output_tokens: int = 1200,
) -> dict[str, Any]:
    """
    Inspect an image under /input or /outputs using a vision-capable OpenAI model.

    Image detail is fixed to OpenAI `original` internally and is not exposed as
    a tool argument.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        normalized, host_path = resolve_readable_virtual_path(path)
        image_detail = "original"
        if not host_path.is_file():
            return {"ok": False, "path": normalized, "error": "not_a_file"}
        if host_path.stat().st_size > 20 * 1024 * 1024:
            return {
                "ok": False,
                "path": normalized,
                "error": "image_too_large",
                "bytes": host_path.stat().st_size,
            }
        mime = image_mime_type(host_path)
        if not mime.startswith("image/"):
            return {"ok": False, "path": normalized, "error": f"unsupported_mime:{mime}"}

        max_output_tokens = max(200, min(max_output_tokens, 4000))
        payload_bytes, sent_mime, vision_transform = prepare_vision_payload_bytes(
            host_path,
            normalized,
            mime,
        )
        data_url = (
            f"data:{sent_mime};base64,"
            + base64.b64encode(payload_bytes).decode("ascii")
        )
        vision_model = active_vision_model_name()
        client = OpenAI()
        response = client.responses.create(
            model=vision_model,
            input=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": (
                                "Inspect this image carefully and answer the question. "
                                "If the image is ambiguous, say so and provide ranked candidates "
                                "with visual evidence instead of overclaiming.\n\n"
                                f"Question: {question}"
                            ),
                        },
                        {
                            "type": "input_image",
                            "image_url": data_url,
                            "detail": image_detail,
                        },
                    ],
                }
            ],
            max_output_tokens=max_output_tokens,
        )
        text = getattr(response, "output_text", "") or ""
        return {
            "ok": True,
            "path": normalized,
            "mime": mime,
            "sent_mime": sent_mime,
            "bytes": host_path.stat().st_size,
            "model": vision_model,
            "detail": image_detail,
            "vision_transform": vision_transform,
            "answer": text,
        }
    except Exception as exc:
        return {"ok": False, "path": path, "error": f"{exc.__class__.__name__}: {exc}"}


@tool
def read_sandbox_file(
    path: str,
    question: str = "",
    max_chars: int = 12000,
    max_rows: int = 20,
    max_cols: int = 12,
    max_pages: int = 5,
    max_output_tokens: int = 1200,
) -> dict[str, Any]:
    """
    Read a sandbox file under /input or /outputs with type-aware handling.

    Text, CSV, JSON, Excel, and PDF files return structured previews. Image files
    return metadata. When `question` is supplied for an image, the tool performs
    a vision read only for profiles that include the `image_inspect` toolset.
    Audio/video and other binaries return metadata only.
    """
    file_info = inspect_sandbox_file.invoke(
        {
            "path": path,
            "max_chars": max_chars,
            "max_rows": max_rows,
            "max_cols": max_cols,
            "max_pages": max_pages,
        }
    )
    if not file_info.get("ok"):
        return file_info

    kind = file_info.get("kind")
    result: dict[str, Any] = {
        "ok": True,
        "path": file_info.get("path", path),
        "kind": kind,
        "read_mode": "typed_preview",
        "file": file_info,
    }
    if kind == "image" and question.strip():
        if ACTIVE_DEEP_AGENT_TOOLSETS and "image_inspect" not in ACTIVE_DEEP_AGENT_TOOLSETS:
            result["question_note"] = (
                "Image vision reads are disabled for the active Deep Agent profile "
                "because it does not include the image_inspect toolset. Use metadata "
                "only or route the task to a profile with image_inspect."
            )
            return result
        vision = inspect_sandbox_image.invoke(
            {
                "path": path,
                "question": question,
                "max_output_tokens": max_output_tokens,
            }
        )
        result["read_mode"] = "image_vision"
        result["vision"] = vision
        result["ok"] = bool(vision.get("ok"))
    elif question.strip():
        result["question_note"] = (
            "The file content was returned as structured preview data. Answer the "
            "question from the preview, or generate a task-specific extractor inside "
            "the sandbox if the preview is insufficient."
        )
    return result


@tool
def crawl_allowed_site(
    start_url: str,
    allowed_domains: list[str] | None = None,
    max_pages: int = 40,
    max_depth: int = 2,
    path_prefixes: list[str] | None = None,
    exclude_url_patterns: list[str] | None = None,
    request_delay_seconds: float = 0.25,
    max_bytes_per_url: int = 2_000_000,
    respect_robots_txt: bool = True,
) -> dict[str, Any]:
    """Crawl a specific public website within an explicit domain/path allowlist.

    This is a controlled site-research tool, not a general web search tool.
    It only fetches http(s) URLs whose final URL remains inside allowed_domains.
    Private, loopback, and local hosts are rejected by default. Results are
    stored under /outputs/_site_crawl/<crawl_id>/ for later search and reading.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        safe_domains = default_allowed_domains_for_url(
            start_url,
            allowed_domains,
            tool_name="site crawl",
        )
        proxy_url = network_tool_proxy_url(safe_domains, purpose="site-crawl")
        policy = CrawlPolicy(
            start_url=start_url,
            allowed_domains=safe_domains,
            max_pages=max_pages,
            max_depth=max_depth,
            path_prefixes=path_prefixes or [],
            exclude_url_patterns=exclude_url_patterns or [],
            request_delay_seconds=request_delay_seconds,
            max_bytes_per_url=max_bytes_per_url,
            respect_robots_txt=respect_robots_txt,
            proxy_url=proxy_url,
        )
        crawl = run_site_crawl(CONFIG.output_dir, policy)
        manifest = crawl["manifest"]
        crawl_id = manifest["crawl_id"]
        virtual_root = f"/outputs/_site_crawl/{crawl_id}"
        return {
            "ok": True,
            "crawl_id": crawl_id,
            "pages_fetched": manifest["pages_fetched"],
            "skipped_count": manifest["skipped_count"],
            "allowed_domains": manifest["allowed_domains"],
            "network_proxy_enabled": bool(proxy_url),
            "virtual_root": virtual_root,
            "manifest_path": f"{virtual_root}/crawl_manifest.json",
            "summary_path": f"{virtual_root}/crawl_summary.md",
            "pages_jsonl_path": f"{virtual_root}/pages.jsonl",
            "sqlite_index_path": f"{virtual_root}/site_index.sqlite",
            "first_pages": crawl["records"][:10],
        }
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "start_url": start_url,
        }


@tool
def extract_allowed_site_links(
    list_url: str,
    allowed_domains: list[str] | None = None,
    path_prefixes: list[str] | None = None,
    required_year: int | None = None,
    required_month: int | None = None,
    date_from: str = "",
    date_to: str = "",
    include_text_patterns: list[str] | None = None,
    exclude_text_patterns: list[str] | None = None,
    include_url_patterns: list[str] | None = None,
    exclude_url_patterns: list[str] | None = None,
    css_selector: str = "",
    allowed_extensions: list[str] | None = None,
    url_contains: str = "",
    max_links: int = 300,
    respect_robots_txt: bool = True,
) -> dict[str, Any]:
    """Extract allowed links from a listing/index page with structured filters.

    Use this before crawling article collections when completeness matters. It
    fetches only the listing URL, parses anchors, filters links by allowed
    domain/path, date, text/URL regex, CSS selector, and extension rules, then
    stores the extracted link set under
    /outputs/_site_crawl/_link_extract/<extract_id>/links.json.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        safe_domains = default_allowed_domains_for_url(
            list_url,
            allowed_domains,
            tool_name="site link extraction",
        )
        proxy_url = network_tool_proxy_url(safe_domains, purpose="site-link-extract")
        policy = LinkExtractPolicy(
            list_url=list_url,
            allowed_domains=safe_domains,
            path_prefixes=path_prefixes or [],
            required_year=required_year,
            required_month=required_month,
            date_from=date_from,
            date_to=date_to,
            include_text_patterns=include_text_patterns or [],
            exclude_text_patterns=exclude_text_patterns or [],
            include_url_patterns=include_url_patterns or [],
            exclude_url_patterns=exclude_url_patterns or [],
            css_selector=css_selector,
            allowed_extensions=allowed_extensions or [],
            url_contains=url_contains,
            max_links=max_links,
            respect_robots_txt=respect_robots_txt,
            proxy_url=proxy_url,
        )
        result = extract_links_from_listing(CONFIG.output_dir, policy)
        manifest = result["manifest"]
        extract_id = manifest["extract_id"]
        virtual_root = f"/outputs/_site_crawl/_link_extract/{extract_id}"
        return {
            "ok": True,
            "extract_id": extract_id,
            "link_count": manifest["link_count"],
            "list_url": manifest["list_url"],
            "required_year": manifest["required_year"],
            "required_month": manifest["required_month"],
            "date_from": manifest["date_from"],
            "date_to": manifest["date_to"],
            "network_proxy_enabled": bool(proxy_url),
            "virtual_root": virtual_root,
            "links_json_path": f"{virtual_root}/links.json",
            "links": manifest["links"][: max(1, min(max_links, 300))],
        }
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "list_url": list_url,
        }


@tool
def crawl_allowed_urls(
    urls: list[str],
    allowed_domains: list[str] | None = None,
    path_prefixes: list[str] | None = None,
    request_delay_seconds: float = 0.25,
    max_bytes_per_url: int = 2_000_000,
    respect_robots_txt: bool = True,
) -> dict[str, Any]:
    """Crawl an explicit list of allowed URLs and build a local site index.

    Use this after extract_allowed_site_links when collection completeness
    matters more than graph traversal order.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        first_url = urls[0] if urls else ""
        safe_domains = default_allowed_domains_for_url(
            first_url,
            allowed_domains,
            tool_name="site URL crawl",
        )
        proxy_url = network_tool_proxy_url(safe_domains, purpose="site-url-crawl")
        policy = CrawlPolicy(
            start_url=first_url,
            allowed_domains=safe_domains,
            max_pages=len(urls),
            max_depth=0,
            path_prefixes=path_prefixes or [],
            request_delay_seconds=request_delay_seconds,
            max_bytes_per_url=max_bytes_per_url,
            respect_robots_txt=respect_robots_txt,
            proxy_url=proxy_url,
        )
        crawl = run_url_crawl(CONFIG.output_dir, urls, policy)
        manifest = crawl["manifest"]
        crawl_id = manifest["crawl_id"]
        virtual_root = f"/outputs/_site_crawl/{crawl_id}"
        return {
            "ok": True,
            "crawl_id": crawl_id,
            "pages_fetched": manifest["pages_fetched"],
            "skipped_count": manifest["skipped_count"],
            "allowed_domains": manifest["allowed_domains"],
            "network_proxy_enabled": bool(proxy_url),
            "virtual_root": virtual_root,
            "manifest_path": f"{virtual_root}/crawl_manifest.json",
            "summary_path": f"{virtual_root}/crawl_summary.md",
            "pages_jsonl_path": f"{virtual_root}/pages.jsonl",
            "sqlite_index_path": f"{virtual_root}/site_index.sqlite",
            "first_pages": crawl["records"][:10],
        }
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "url_count": len(urls),
        }


@tool
def search_houjin_bangou_by_name(
    query: str,
    match_type: str = "prefix",
    include_closed: bool = True,
    max_results: int = 20,
    try_name_variants: bool = True,
    respect_robots_txt: bool = True,
) -> dict[str, Any]:
    """Search Japan's official Corporate Number Publication Site by company name.

    Use this for tasks that ask whether a Japanese corporation exists on the
    National Tax Agency Corporate Number Publication Site. Prefer
    match_type="prefix" for full legal names, and use "partial" for broader
    candidate discovery. If try_name_variants is true and the exact query has
    no hits, the tool may retry common normalized variants such as removing a
    leading legal designator, while still judging exact_matches against the
    original query. The result includes exact_matches, candidate rows, search
    attempts, and saved audit artifacts under
    /outputs/_official_search/houjin_bangou/.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        proxy_url = network_tool_proxy_url(
            ["www.houjin-bangou.nta.go.jp"],
            purpose="houjin-bangou-search",
        )
        result = run_houjin_bangou_search(
            CONFIG.output_dir,
            HoujinBangouSearchPolicy(
                query=query,
                match_type=match_type,
                include_closed=include_closed,
                max_results=max_results,
                try_name_variants=try_name_variants,
                respect_robots_txt=respect_robots_txt,
                proxy_url=proxy_url,
            ),
        )
        result["network_proxy_enabled"] = bool(proxy_url)
        return result
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "query": query,
            "match_type": match_type,
        }


def browser_tool_run_id(prefix: str, task: str, allowed_domains: list[str]) -> str:
    digest = hashlib.sha256(
        (task + "|" + "|".join(sorted(allowed_domains))).encode("utf-8")
    ).hexdigest()[:10]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    safe_prefix = re.sub(r"[^a-zA-Z0-9_-]+", "_", prefix.strip() or "browser").strip("_")
    return f"{safe_prefix}_{timestamp}_{digest}"


def clamp_playwright_delay_ms(value: int | None, *, default_ms: int, max_ms: int) -> int:
    if value is None:
        return default_ms
    try:
        delay_ms = int(value)
    except (TypeError, ValueError):
        return default_ms
    return max(0, min(delay_ms, max_ms))


def playwright_domain_delay_key(allowed_domains: list[str]) -> str:
    return "|".join(sorted(normalize_domain(domain) for domain in allowed_domains))


def apply_playwright_domain_delay(allowed_domains: list[str], min_delay_ms: int) -> int:
    if min_delay_ms <= 0:
        return 0
    import time

    key = playwright_domain_delay_key(allowed_domains)
    last_finish = PLAYWRIGHT_DOMAIN_LAST_FINISH.get(key)
    if last_finish is None:
        return 0
    elapsed_ms = int((time.monotonic() - last_finish) * 1000)
    wait_ms = max(0, min_delay_ms - elapsed_ms)
    if wait_ms > 0:
        time.sleep(wait_ms / 1000)
    return wait_ms


def record_playwright_domain_finish(allowed_domains: list[str]) -> None:
    if not allowed_domains:
        return
    import time

    PLAYWRIGHT_DOMAIN_LAST_FINISH[playwright_domain_delay_key(allowed_domains)] = time.monotonic()


def validate_public_allowed_domains(
    allowed_domains: list[str] | None,
    *,
    tool_name: str,
) -> list[str]:
    if not allowed_domains:
        raise ValueError(f"allowed_domains is required for {tool_name} tasks.")
    normalized: list[str] = []
    for item in allowed_domains:
        raw = (item or "").strip()
        if not raw:
            continue
        if "*" in raw and not raw.startswith("*.") and not raw.startswith("http*://"):
            raise ValueError(f"Unsupported allowed domain wildcard: {raw}")
        domain_part = raw
        if "://" in domain_part:
            parsed = urllib.parse.urlsplit(domain_part.replace("http*://", "https://", 1))
            domain_part = parsed.hostname or ""
        elif domain_part.startswith("*."):
            domain_part = domain_part[2:]
        if not domain_part:
            raise ValueError(f"Invalid allowed domain: {raw}")
        normalized_domain = normalize_domain(domain_part)
        if is_private_or_local_host(normalized_domain):
            raise ValueError(f"Private/local domains are not allowed for {tool_name}: {raw}")
        normalized.append(raw)
    if not normalized:
        raise ValueError("allowed_domains is empty after normalization.")
    return normalized


def default_allowed_domains_for_url(
    url: str,
    allowed_domains: list[str] | None,
    *,
    tool_name: str,
) -> list[str]:
    if allowed_domains:
        return validate_public_allowed_domains(allowed_domains, tool_name=tool_name)
    parsed = urllib.parse.urlsplit(url if "://" in url else f"https://{url}")
    host = parsed.hostname or ""
    return validate_public_allowed_domains([host], tool_name=tool_name)


def network_tool_proxy_url(allowed_domains: list[str], *, purpose: str) -> str:
    if CONFIG is None or not CONFIG.egress_proxy_url:
        return ""
    if not CONFIG.egress_proxy_signing_secret:
        raise ValueError(
            "EGRESS_PROXY_URL is configured but EGRESS_PROXY_SIGNING_SECRET is missing."
        )
    token = create_egress_token(
        allowed_domains=allowed_domains,
        secret=CONFIG.egress_proxy_signing_secret,
        purpose=purpose,
        ttl_seconds=900,
    )
    parsed = urllib.parse.urlsplit(CONFIG.egress_proxy_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError(f"Invalid EGRESS_PROXY_URL: {CONFIG.egress_proxy_url}")
    username = urllib.parse.quote(token, safe="")
    password = urllib.parse.quote("x", safe="")
    netloc = f"{username}:{password}@{parsed.hostname}"
    if parsed.port is not None:
        netloc += f":{parsed.port}"
    return urllib.parse.urlunsplit((parsed.scheme, netloc, parsed.path or "", "", ""))


def playwright_proxy_settings(proxy_url: str) -> dict[str, str] | None:
    if not proxy_url:
        return None
    parsed = urllib.parse.urlsplit(proxy_url)
    host = parsed.hostname or ""
    if not host:
        raise ValueError("proxy URL host is required")
    netloc = host
    if parsed.port is not None:
        netloc += f":{parsed.port}"
    settings: dict[str, str] = {
        "server": urllib.parse.urlunsplit((parsed.scheme or "http", netloc, "", "", "")),
    }
    if parsed.username is not None:
        settings["username"] = urllib.parse.unquote(parsed.username)
    if parsed.password is not None:
        settings["password"] = urllib.parse.unquote(parsed.password)
    return settings


def allowed_domain_matches(hostname: str, allowed_domain: str) -> bool:
    raw = allowed_domain.strip().lower()
    if "://" in raw:
        parsed = urllib.parse.urlsplit(raw.replace("http*://", "https://", 1))
        raw = parsed.hostname or ""
    wildcard = raw.startswith("*.")
    if wildcard:
        raw = raw[2:]
    host = normalize_domain(hostname)
    domain = normalize_domain(raw)
    if wildcard:
        return host == domain or host.endswith("." + domain)
    return host == domain


def is_url_allowed_by_domains(url: str, allowed_domains: list[str]) -> bool:
    parsed = urllib.parse.urlsplit(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    hostname = parsed.hostname or ""
    if not hostname or is_private_or_local_host(normalize_domain(hostname)):
        return False
    return any(allowed_domain_matches(hostname, domain) for domain in allowed_domains)


def normalize_playwright_start_url(start_url: str, allowed_domains: list[str]) -> str:
    url = (start_url or "").strip()
    if not url:
        raise ValueError("start_url is required.")
    if "://" not in url:
        url = "https://" + url
    parsed = urllib.parse.urlsplit(url)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError(f"Only http/https URLs are supported: {start_url}")
    if not is_url_allowed_by_domains(url, allowed_domains):
        raise ValueError(f"start_url is outside allowed_domains: {start_url}")
    return url


def limited_jsonable(value: Any, *, max_items: int = 30, max_chars: int = 4000) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str):
        return value[:max_chars]
    if isinstance(value, (list, tuple)):
        return [limited_jsonable(item, max_items=max_items, max_chars=max_chars) for item in value[:max_items]]
    if isinstance(value, dict):
        result: dict[str, Any] = {}
        for index, (key, item) in enumerate(value.items()):
            if index >= max_items:
                break
            result[str(key)] = limited_jsonable(item, max_items=max_items, max_chars=max_chars)
        return result
    return str(value)[:max_chars]


def compact_text(value: str, *, max_chars: int = 12000) -> str:
    return re.sub(r"\s+", " ", value or "").strip()[:max_chars]


PLAYWRIGHT_MAX_FILL_CHARS = 120
PLAYWRIGHT_MAX_SELECT_CHARS = 120
PLAYWRIGHT_MAX_URL_CHARS = 2000
PLAYWRIGHT_MAX_URL_QUERY_CHARS = 600
PLAYWRIGHT_MAX_URL_VALUE_CHARS = 180
PLAYWRIGHT_MAX_POST_CHARS = 5000
PLAYWRIGHT_SAFE_KEYS = {
    "Enter",
    "Tab",
    "Escape",
    "Backspace",
    "Delete",
    "Space",
    "ArrowDown",
    "ArrowLeft",
    "ArrowRight",
    "ArrowUp",
    "Home",
    "End",
    "PageDown",
    "PageUp",
}
PLAYWRIGHT_SECRET_PATTERNS = [
    re.compile(pattern, re.I)
    for pattern in (
        r"-----BEGIN [A-Z ]{0,40}(PRIVATE KEY|SECRET|TOKEN)",
        r"\b(sk|sk-proj|ghp|github_pat|xox[baprs])-[-_A-Za-z0-9]{20,}",
        r"\b(AKIA|ASIA)[A-Z0-9]{16}\b",
        r"\b(openai|api|access|refresh|secret|private|password|passwd|token|credential)[-_ ]?(key|token|secret|password)?\b\s*[:=]",
        r"\b(input|outputs|workspace)/[^ \t\r\n]+",
        r"[A-Za-z]:\\[^ \t\r\n]+",
    )
]


def shannon_entropy(value: str) -> float:
    if not value:
        return 0.0
    counts: dict[str, int] = {}
    for char in value:
        counts[char] = counts.get(char, 0) + 1
    total = len(value)
    return -sum((count / total) * math.log2(count / total) for count in counts.values())


def looks_like_encoded_secret(value: str) -> bool:
    compact = re.sub(r"\s+", "", value)
    if len(compact) < 32:
        return False
    if re.fullmatch(r"[A-Fa-f0-9]{48,}", compact):
        return True
    if re.fullmatch(r"[A-Za-z0-9+/_=-]{48,}", compact) and shannon_entropy(compact) >= 4.4:
        return True
    if len(compact) >= 48 and shannon_entropy(compact) >= 4.8:
        return True
    return False


def validate_playwright_egress_text(
    value: Any,
    *,
    context: str,
    max_chars: int,
    allow_multiline: bool = False,
) -> str:
    text = "" if value is None else str(value)
    if len(text) > max_chars:
        raise ValueError(
            f"Egress guard rejected {context}: value length {len(text)} exceeds {max_chars}."
        )
    if not allow_multiline and ("\n" in text or "\r" in text):
        raise ValueError(f"Egress guard rejected {context}: multiline values are not allowed.")
    if any(ord(char) < 32 and char not in "\t\n\r" for char in text):
        raise ValueError(f"Egress guard rejected {context}: control characters are not allowed.")
    stripped = text.strip()
    if len(stripped) > 40 and stripped[0:1] in {"{", "[", "<"}:
        raise ValueError(f"Egress guard rejected {context}: structured payload-like value.")
    for pattern in PLAYWRIGHT_SECRET_PATTERNS:
        if pattern.search(text):
            raise ValueError(f"Egress guard rejected {context}: secret/path-like value.")
    if looks_like_encoded_secret(text):
        raise ValueError(f"Egress guard rejected {context}: encoded/high-entropy value.")
    return text


def validate_playwright_url_egress(url: str, *, context: str) -> None:
    if len(url) > PLAYWRIGHT_MAX_URL_CHARS:
        raise ValueError(
            f"Egress guard rejected {context}: URL length {len(url)} exceeds {PLAYWRIGHT_MAX_URL_CHARS}."
        )
    parsed = urllib.parse.urlsplit(url)
    query = parsed.query or ""
    fragment = parsed.fragment or ""
    if len(query) > PLAYWRIGHT_MAX_URL_QUERY_CHARS:
        raise ValueError(
            f"Egress guard rejected {context}: query length {len(query)} exceeds {PLAYWRIGHT_MAX_URL_QUERY_CHARS}."
        )
    if len(fragment) > PLAYWRIGHT_MAX_URL_VALUE_CHARS:
        raise ValueError(
            f"Egress guard rejected {context}: fragment length {len(fragment)} exceeds {PLAYWRIGHT_MAX_URL_VALUE_CHARS}."
        )
    if fragment:
        validate_playwright_egress_text(
            fragment,
            context=f"{context} fragment",
            max_chars=PLAYWRIGHT_MAX_URL_VALUE_CHARS,
        )
    for segment in parsed.path.split("/"):
        if not segment:
            continue
        validate_playwright_egress_text(
            urllib.parse.unquote_plus(segment),
            context=f"{context} path segment",
            max_chars=PLAYWRIGHT_MAX_URL_VALUE_CHARS,
        )
    for key, value in urllib.parse.parse_qsl(query, keep_blank_values=True):
        validate_playwright_egress_text(
            key,
            context=f"{context} query key",
            max_chars=80,
        )
        validate_playwright_egress_text(
            value,
            context=f"{context} query value",
            max_chars=PLAYWRIGHT_MAX_URL_VALUE_CHARS,
        )


def step_fill_value(step: dict[str, Any]) -> Any:
    value_source = step.get("value")
    if value_source is None and step.get("input_text") is not None:
        value_source = step.get("input_text")
    if value_source is None and any(
        step.get(key_name)
        for key_name in (
            "selector",
            "role",
            "label",
            "placeholder",
            "alt_text",
            "title",
        )
    ):
        value_source = step.get("text")
    return value_source


def validate_playwright_step_egress(
    step: dict[str, Any],
    *,
    allowed_domains: list[str],
    step_index: int,
) -> None:
    action = str(step.get("action") or "").strip().lower()
    if action == "goto":
        url = normalize_playwright_start_url(str(step.get("url") or ""), allowed_domains)
        validate_playwright_url_egress(url, context=f"step {step_index} goto")
    elif action == "fill":
        validate_playwright_egress_text(
            step_fill_value(step),
            context=f"step {step_index} fill",
            max_chars=PLAYWRIGHT_MAX_FILL_CHARS,
        )
    elif action == "select":
        for key in ("value", "option_label"):
            if step.get(key) is not None:
                validate_playwright_egress_text(
                    step.get(key),
                    context=f"step {step_index} select {key}",
                    max_chars=PLAYWRIGHT_MAX_SELECT_CHARS,
                )
    elif action == "press":
        key = str(step.get("key") or "").strip()
        if key and key not in PLAYWRIGHT_SAFE_KEYS:
            raise ValueError(f"Egress guard rejected step {step_index}: unsupported key {key!r}.")


def validate_playwright_steps_egress(
    steps: list[dict[str, Any]],
    *,
    allowed_domains: list[str],
) -> None:
    for index, step in enumerate(steps, start=1):
        validate_playwright_step_egress(
            dict(step or {}),
            allowed_domains=allowed_domains,
            step_index=index,
        )


def playwright_locator(page: Any, target: dict[str, Any]) -> Any:
    selector = str(target.get("selector") or "").strip()
    if selector:
        return page.locator(selector).first
    role = str(target.get("role") or "").strip()
    name = str(target.get("name") or "").strip()
    if role and name:
        return page.get_by_role(role, name=name, exact=False).first
    label = str(target.get("label") or "").strip()
    if label:
        return page.get_by_label(label, exact=False).first
    placeholder = str(target.get("placeholder") or "").strip()
    if placeholder:
        return page.get_by_placeholder(placeholder, exact=False).first
    text = str(target.get("text") or "").strip()
    if text:
        return page.get_by_text(text, exact=False).first
    alt_text = str(target.get("alt_text") or "").strip()
    if alt_text:
        return page.get_by_alt_text(alt_text, exact=False).first
    title = str(target.get("title") or "").strip()
    if title:
        return page.get_by_title(title, exact=False).first
    raise ValueError("Action target requires selector, role+name, label, placeholder, text, alt_text, or title.")


def playwright_extract_page_state(page: Any, *, max_text_chars: int = 12000) -> dict[str, Any]:
    def eval_all(selector: str, script: str, *, timeout_ms: int = 3000) -> Any:
        try:
            page.locator(selector).first.wait_for(state="attached", timeout=timeout_ms)
        except Exception:
            pass
        try:
            return limited_jsonable(page.locator(selector).evaluate_all(script), max_items=100)
        except Exception as exc:
            return {"error": f"{exc.__class__.__name__}: {exc}"}

    try:
        body_text = page.locator("body").inner_text(timeout=4000)
    except Exception:
        body_text = ""

    return {
        "url": page.url,
        "title": page.title(),
        "text_preview": compact_text(body_text, max_chars=max_text_chars),
        "links": eval_all(
            "a",
            """
            els => els.slice(0, 120).map(a => ({
              text: (a.innerText || a.textContent || '').trim().slice(0, 180),
              href: a.href || a.getAttribute('href') || '',
              title: a.getAttribute('title') || ''
            }))
            """,
        ),
        "inputs": eval_all(
            "input, textarea, select",
            """
            els => els.slice(0, 120).map(el => ({
              tag: el.tagName.toLowerCase(),
              type: el.getAttribute('type') || '',
              name: el.getAttribute('name') || '',
              id: el.getAttribute('id') || '',
              placeholder: el.getAttribute('placeholder') || '',
              aria_label: el.getAttribute('aria-label') || '',
              value: el.value || '',
              label: (() => {
                if (el.labels && el.labels.length) {
                  return Array.from(el.labels).map(l => l.innerText || l.textContent || '').join(' ').trim();
                }
                return '';
              })()
            }))
            """,
        ),
        "buttons": eval_all(
            "button, input[type=button], input[type=submit], input[type=reset], a[role=button]",
            """
            els => els.slice(0, 80).map(el => ({
              tag: el.tagName.toLowerCase(),
              type: el.getAttribute('type') || '',
              name: el.getAttribute('name') || '',
              id: el.getAttribute('id') || '',
              text: (el.innerText || el.value || el.textContent || '').trim().slice(0, 180),
              aria_label: el.getAttribute('aria-label') || ''
            }))
            """,
        ),
        "forms": eval_all(
            "form",
            """
            els => els.slice(0, 20).map(form => ({
              method: form.getAttribute('method') || 'get',
              action: form.action || form.getAttribute('action') || '',
              text: (form.innerText || form.textContent || '').trim().slice(0, 500),
              inputs: Array.from(form.querySelectorAll('input, textarea, select')).slice(0, 60).map(el => ({
                tag: el.tagName.toLowerCase(),
                type: el.getAttribute('type') || '',
                name: el.getAttribute('name') || '',
                id: el.getAttribute('id') || '',
                placeholder: el.getAttribute('placeholder') || '',
                value: el.value || ''
              })),
              buttons: Array.from(form.querySelectorAll('button, input[type=submit], input[type=button]')).slice(0, 30).map(el => ({
                tag: el.tagName.toLowerCase(),
                type: el.getAttribute('type') || '',
                text: (el.innerText || el.value || el.textContent || '').trim().slice(0, 180)
              }))
            }))
            """,
        ),
        "tables": eval_all(
            "table",
            """
            els => els.slice(0, 20).map(table => Array.from(table.querySelectorAll('tr')).slice(0, 12).map(row =>
              Array.from(row.querySelectorAll('th,td')).slice(0, 12).map(cell =>
                (cell.innerText || cell.textContent || '').trim().slice(0, 160)
              )
            ))
            """,
        ),
    }


def playwright_page_shows_rate_limit(page: Any) -> bool:
    try:
        title = str(page.title() or "")
    except Exception:
        title = ""
    try:
        body = str(page.locator("body").inner_text(timeout=2000) or "")
    except Exception:
        body = ""
    probe = f"{title}\n{body[:1000]}".lower()
    return "too many requests" in probe or bool(re.search(r"\b429\b", probe))


def run_playwright_steps(
    *,
    start_url: str,
    allowed_domains: list[str],
    steps: list[dict[str, Any]],
    output_dir: Path,
    timeout_seconds: int,
    capture_screenshot: bool,
    max_text_chars: int,
    min_action_delay_ms: int,
    rate_limit_backoff_ms: int,
    proxy_url: str = "",
) -> dict[str, Any]:
    import time

    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        raise RuntimeError(
            "playwright is not installed in the runner environment. "
            "Install playwright and rebuild the agent image."
        ) from exc

    action_log: list[dict[str, Any]] = []
    blocked_urls: list[str] = []
    screenshots: list[str] = []
    rate_limit_events: list[dict[str, Any]] = []
    started = time.monotonic()
    safe_timeout_ms = max(10, min(timeout_seconds, 240)) * 1000

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context_kwargs: dict[str, Any] = {
            "locale": "ja-JP",
            "timezone_id": "Asia/Tokyo",
            "viewport": {"width": 1365, "height": 900},
        }
        proxy_settings = playwright_proxy_settings(proxy_url)
        if proxy_settings:
            context_kwargs["proxy"] = proxy_settings
        context = browser.new_context(**context_kwargs)
        page = context.new_page()
        page.set_default_timeout(min(safe_timeout_ms, 30000))

        def route_handler(route: Any) -> None:
            request_url = route.request.url
            if request_url.startswith(("data:", "blob:", "about:")):
                route.continue_()
                return
            if is_url_allowed_by_domains(request_url, allowed_domains):
                try:
                    validate_playwright_url_egress(
                        request_url,
                        context=f"{route.request.method} request",
                    )
                    method = str(route.request.method or "GET").upper()
                    if method not in {"GET", "HEAD", "OPTIONS"}:
                        post_data = route.request.post_data or ""
                        if len(post_data) > PLAYWRIGHT_MAX_POST_CHARS:
                            raise ValueError(
                                f"POST body length {len(post_data)} exceeds {PLAYWRIGHT_MAX_POST_CHARS}."
                            )
                except Exception as exc:
                    blocked_urls.append(f"{request_url} [egress_guard: {exc}]")
                    route.abort()
                    return
                route.continue_()
                return
            blocked_urls.append(request_url)
            route.abort()

        page.route("**/*", route_handler)

        def apply_action_delay(log_item: dict[str, Any]) -> None:
            if min_action_delay_ms <= 0:
                return
            page.wait_for_timeout(min_action_delay_ms)
            log_item["politeness_delay_ms"] = min_action_delay_ms

        def apply_rate_limit_backoff(log_item: dict[str, Any]) -> None:
            status = log_item.get("http_status")
            detected = status == 429 or playwright_page_shows_rate_limit(page)
            if not detected:
                return
            event = {
                "step": log_item.get("step", 0),
                "action": log_item.get("action", ""),
                "url": page.url,
                "http_status": status,
                "backoff_ms": rate_limit_backoff_ms,
            }
            rate_limit_events.append(event)
            log_item["rate_limit_detected"] = True
            log_item["rate_limit_backoff_ms"] = rate_limit_backoff_ms
            if rate_limit_backoff_ms > 0:
                page.wait_for_timeout(rate_limit_backoff_ms)

        try:
            response = page.goto(start_url, wait_until="domcontentloaded", timeout=safe_timeout_ms)
            try:
                page.wait_for_load_state("networkidle", timeout=5000)
            except Exception:
                pass
            initial_log: dict[str, Any] = {"action": "goto", "url": start_url, "ok": True}
            if response is not None:
                initial_log["http_status"] = response.status
            apply_rate_limit_backoff(initial_log)
            apply_action_delay(initial_log)
            action_log.append(initial_log)

            for index, raw_step in enumerate(steps[:20], start=1):
                step = dict(raw_step or {})
                action = str(step.get("action") or "").strip().lower()
                if not action:
                    raise ValueError(f"Step {index} missing action.")
                log_item: dict[str, Any] = {"step": index, "action": action}
                if action == "goto":
                    url = normalize_playwright_start_url(str(step.get("url") or ""), allowed_domains)
                    response = page.goto(url, wait_until="domcontentloaded", timeout=safe_timeout_ms)
                    log_item["url"] = url
                    if response is not None:
                        log_item["http_status"] = response.status
                elif action == "click":
                    locator = playwright_locator(page, step)
                    locator.click(timeout=min(safe_timeout_ms, 15000))
                    log_item["target"] = limited_jsonable(step)
                elif action == "fill":
                    locator = playwright_locator(page, step)
                    value = validate_playwright_egress_text(
                        step_fill_value(step),
                        context=f"step {index} fill",
                        max_chars=PLAYWRIGHT_MAX_FILL_CHARS,
                    )
                    locator.fill(value, timeout=min(safe_timeout_ms, 15000))
                    log_item["target"] = limited_jsonable({k: v for k, v in step.items() if k != "value"})
                    log_item["value_length"] = len(value)
                elif action == "press":
                    key = str(step.get("key") or "").strip()
                    if not key:
                        raise ValueError(f"Step {index} press action requires key.")
                    if any(step.get(key_name) for key_name in ("selector", "role", "label", "placeholder", "text", "alt_text", "title")):
                        playwright_locator(page, step).press(key, timeout=min(safe_timeout_ms, 15000))
                    else:
                        page.keyboard.press(key)
                    log_item["key"] = key
                elif action == "select":
                    locator = playwright_locator(page, step)
                    value = step.get("value")
                    label = step.get("option_label")
                    index_value = step.get("option_index")
                    if label is not None:
                        selected = locator.select_option(label=str(label), timeout=min(safe_timeout_ms, 15000))
                    elif index_value is not None:
                        selected = locator.select_option(index=int(index_value), timeout=min(safe_timeout_ms, 15000))
                    else:
                        selected = locator.select_option(value=str(value), timeout=min(safe_timeout_ms, 15000))
                    log_item["selected"] = limited_jsonable(selected)
                elif action == "wait":
                    milliseconds = int(step.get("milliseconds") or 1000)
                    page.wait_for_timeout(max(0, min(milliseconds, 10000)))
                    log_item["milliseconds"] = milliseconds
                elif action == "wait_for_selector":
                    selector = str(step.get("selector") or "").strip()
                    if not selector:
                        raise ValueError(f"Step {index} wait_for_selector requires selector.")
                    page.locator(selector).first.wait_for(timeout=min(safe_timeout_ms, 15000))
                    log_item["selector"] = selector
                elif action == "wait_for_text":
                    text = str(step.get("text") or "").strip()
                    if not text:
                        raise ValueError(f"Step {index} wait_for_text requires text.")
                    page.get_by_text(text, exact=False).first.wait_for(timeout=min(safe_timeout_ms, 15000))
                    log_item["text"] = text
                elif action == "screenshot":
                    screenshot_name = f"screenshot_{index:02d}.png"
                    page.screenshot(path=str(output_dir / screenshot_name), full_page=bool(step.get("full_page", True)))
                    screenshots.append(screenshot_name)
                    log_item["screenshot"] = screenshot_name
                elif action == "extract_text":
                    if any(step.get(key_name) for key_name in ("selector", "role", "label", "placeholder", "text", "alt_text", "title")):
                        extracted = playwright_locator(page, step).inner_text(timeout=min(safe_timeout_ms, 15000))
                    else:
                        extracted = page.locator("body").inner_text(timeout=min(safe_timeout_ms, 15000))
                    log_item["text_preview"] = compact_text(extracted, max_chars=4000)
                else:
                    raise ValueError(f"Unsupported Playwright action: {action}")

                try:
                    page.wait_for_load_state("domcontentloaded", timeout=3000)
                except Exception:
                    pass
                try:
                    page.wait_for_load_state("networkidle", timeout=5000)
                except Exception:
                    pass
                apply_rate_limit_backoff(log_item)
                if action != "wait":
                    apply_action_delay(log_item)
                log_item["ok"] = True
                log_item["url"] = page.url
                action_log.append(log_item)

            if capture_screenshot:
                screenshot_name = "final.png"
                page.screenshot(path=str(output_dir / screenshot_name), full_page=True)
                screenshots.append(screenshot_name)

            page_state = playwright_extract_page_state(page, max_text_chars=max_text_chars)
        finally:
            context.close()
            browser.close()

    return {
        "elapsed_seconds": round(time.monotonic() - started, 3),
        "actions": action_log,
        "blocked_urls": blocked_urls[:200],
        "blocked_url_count": len(blocked_urls),
        "rate_limit_events": rate_limit_events[:50],
        "rate_limit_event_count": len(rate_limit_events),
        "screenshots": screenshots,
        "page": page_state,
    }


@tool
def run_playwright_task(
    task: str,
    start_url: str,
    allowed_domains: list[str],
    steps: list[dict[str, Any]] | None = None,
    timeout_seconds: int = 120,
    capture_screenshot: bool = False,
    max_text_chars: int = 12000,
    min_action_delay_ms: int = PLAYWRIGHT_DEFAULT_MIN_ACTION_DELAY_MS,
    rate_limit_backoff_ms: int = PLAYWRIGHT_DEFAULT_RATE_LIMIT_BACKOFF_MS,
) -> dict[str, Any]:
    """Run a deterministic Playwright browser task with restricted actions.

    Use this for interactive public sites. It is not arbitrary code execution
    and it is not broad web search. Provide a start_url and explicit
    allowed_domains. Optional steps may include:
    goto, click, fill, press, select, wait, wait_for_selector, wait_for_text,
    extract_text, and screenshot. Targets can use selector, role+name, label,
    placeholder, text, alt_text, or title. For fill actions, pass value or
    input_text; text is also accepted as the fill value when another target
    field such as selector or placeholder is present. Egress guard rejects long
    values, secret/path-like strings, high-entropy encoded payloads, structured
    payload-like values, and long URL query/fragment values. File upload actions
    are not supported.

    Generic rate-limit avoidance is enabled by default: min_action_delay_ms
    defaults to 1000 and is applied after browser actions and between browser
    tool calls for the same allowed domain set. If a page appears to be rate
    limited (HTTP 429 or "Too Many Requests"), rate_limit_backoff_ms is applied
    before returning the observed page state. Results are saved under
    /outputs/_playwright/<run_id>/result.json and summary.md.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    run_id = ""
    safe_domains: list[str] = []
    try:
        safe_domains = validate_public_allowed_domains(allowed_domains, tool_name="Playwright")
        safe_start_url = normalize_playwright_start_url(start_url, safe_domains)
        validate_playwright_url_egress(safe_start_url, context="start_url")
        limited_steps = list(steps or [])[:20]
        validate_playwright_steps_egress(limited_steps, allowed_domains=safe_domains)
        safe_min_action_delay_ms = clamp_playwright_delay_ms(
            min_action_delay_ms,
            default_ms=PLAYWRIGHT_DEFAULT_MIN_ACTION_DELAY_MS,
            max_ms=PLAYWRIGHT_MAX_MIN_ACTION_DELAY_MS,
        )
        safe_rate_limit_backoff_ms = clamp_playwright_delay_ms(
            rate_limit_backoff_ms,
            default_ms=PLAYWRIGHT_DEFAULT_RATE_LIMIT_BACKOFF_MS,
            max_ms=PLAYWRIGHT_MAX_RATE_LIMIT_BACKOFF_MS,
        )
        proxy_url = network_tool_proxy_url(safe_domains, purpose="playwright")
        domain_wait_ms = apply_playwright_domain_delay(safe_domains, safe_min_action_delay_ms)
        run_id = browser_tool_run_id("playwright", task + "|" + safe_start_url, safe_domains)
        playwright_dir = CONFIG.output_dir / "_playwright" / run_id
        playwright_dir.mkdir(parents=True, exist_ok=True)
        try:
            result = run_playwright_steps(
                start_url=safe_start_url,
                allowed_domains=safe_domains,
                steps=limited_steps,
                output_dir=playwright_dir,
                timeout_seconds=timeout_seconds,
                capture_screenshot=capture_screenshot,
                max_text_chars=max(1000, min(max_text_chars, 30000)),
                min_action_delay_ms=safe_min_action_delay_ms,
                rate_limit_backoff_ms=safe_rate_limit_backoff_ms,
                proxy_url=proxy_url,
            )
        finally:
            record_playwright_domain_finish(safe_domains)
        virtual_root = f"/outputs/_playwright/{run_id}"
        saved_result: dict[str, Any] = {
            "ok": True,
            "run_id": run_id,
            "task": task,
            "start_url": safe_start_url,
            "allowed_domains": safe_domains,
            "timeout_seconds": timeout_seconds,
            "capture_screenshot": capture_screenshot,
            "min_action_delay_ms": safe_min_action_delay_ms,
            "rate_limit_backoff_ms": safe_rate_limit_backoff_ms,
            "domain_wait_ms": domain_wait_ms,
            "network_proxy_enabled": bool(proxy_url),
            "virtual_root": virtual_root,
            **result,
        }
        (playwright_dir / "result.json").write_text(
            json.dumps(saved_result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        page = saved_result.get("page") or {}
        summary = [
            "# Playwright Task Summary",
            "",
            f"- Run ID: {run_id}",
            f"- Start URL: {safe_start_url}",
            f"- Current URL: {page.get('url') or ''}",
            f"- Title: {page.get('title') or ''}",
            f"- Allowed domains: {', '.join(safe_domains)}",
            f"- Elapsed seconds: {saved_result.get('elapsed_seconds')}",
            f"- Actions: {len(saved_result.get('actions') or [])}",
            f"- Blocked external requests: {saved_result.get('blocked_url_count')}",
            f"- Min action delay ms: {saved_result.get('min_action_delay_ms')}",
            f"- Domain wait ms before start: {saved_result.get('domain_wait_ms')}",
            f"- Network proxy enabled: {saved_result.get('network_proxy_enabled')}",
            f"- Rate-limit events: {saved_result.get('rate_limit_event_count')}",
            "",
            "## Text Preview",
            "",
            str(page.get("text_preview") or "")[:12000],
        ]
        (playwright_dir / "summary.md").write_text("\n".join(summary), encoding="utf-8")
        return {
            **saved_result,
            "result_json_path": f"{virtual_root}/result.json",
            "summary_path": f"{virtual_root}/summary.md",
            "screenshot_paths": [
                f"{virtual_root}/{name}" for name in saved_result.get("screenshots", [])
            ],
        }
    except Exception as exc:
        if CONFIG is not None and run_id:
            error_dir = CONFIG.output_dir / "_playwright" / run_id
            error_dir.mkdir(parents=True, exist_ok=True)
            (error_dir / "error.json").write_text(
                json.dumps(
                    {
                        "ok": False,
                        "error": f"{exc.__class__.__name__}: {exc}",
                        "task": task,
                        "start_url": start_url,
                        "allowed_domains": allowed_domains,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "task": task[:1000],
            "start_url": start_url,
            "allowed_domains": allowed_domains,
        }


@tool
def search_site_crawl(
    query: str,
    crawl_id: str = "",
    max_results: int = 8,
) -> dict[str, Any]:
    """Search a previously created site crawl by keyword and return sourced snippets."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        result = search_site_crawl_index(
            CONFIG.output_dir,
            query=query,
            crawl_id=crawl_id or None,
            max_results=max_results,
        )
        for item in result["results"]:
            item["virtual_path"] = f"/outputs/_site_crawl/{result['crawl_id']}/{item['text_path']}"
        return {"ok": True, **result}
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "query": query,
            "crawl_id": crawl_id,
        }


@tool
def read_crawled_page(
    page: str,
    crawl_id: str = "",
    max_chars: int = 12000,
) -> dict[str, Any]:
    """Read one page from a site crawl by page_id or URL."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        result = read_site_crawl_page(
            CONFIG.output_dir,
            page=page,
            crawl_id=crawl_id or None,
            max_chars=max(1000, min(max_chars, 50000)),
        )
        result["virtual_path"] = (
            f"/outputs/_site_crawl/{result['crawl_id']}/{result['page']['text_path']}"
        )
        return {"ok": True, **result}
    except Exception as exc:
        return {
            "ok": False,
            "error": f"{exc.__class__.__name__}: {exc}",
            "page": page,
            "crawl_id": crawl_id,
        }


@tool
def list_site_crawls(max_crawls: int = 10) -> dict[str, Any]:
    """List site crawls created during this run."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        result = list_site_crawl_runs(CONFIG.output_dir, max_crawls=max_crawls)
        for item in result["crawls"]:
            item["virtual_root"] = f"/outputs/_site_crawl/{item['crawl_id']}"
        return {"ok": True, **result}
    except Exception as exc:
        return {"ok": False, "error": f"{exc.__class__.__name__}: {exc}"}


def normalize_export_path(path: str | Path) -> tuple[str, Path]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    raw = str(path).replace("\\", "/")
    normalized = posixpath.normpath(raw)
    if normalized.startswith("/outputs/"):
        rel = normalized.removeprefix("/outputs/")
    elif normalized.startswith("/exports/"):
        rel = normalized.removeprefix("/exports/")
    else:
        raise ValueError(f"Exported artifact path must be under /outputs or /exports: {path}")
    if rel in {"", "."} or rel == ".." or rel.startswith("../"):
        raise ValueError(f"Invalid exported artifact path: {path}")
    host_path = (CONFIG.clean_export_dir / rel).resolve()
    clean_root = CONFIG.clean_export_dir.resolve()
    if not (host_path == clean_root or is_relative_to(host_path, clean_root)):
        raise ValueError(f"Resolved export path escaped clean export root: {path}")
    return "/exports/" + rel, host_path


def call_controller_gate(artifacts: list[str]) -> dict[str, Any]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    payload = {
        "artifacts": artifacts,
        "xlsx_dangerous_formula_action": CONFIG.xlsx_dangerous_formula_action,
    }
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if CONFIG.sandbox_controller_token:
        headers["Authorization"] = f"Bearer {CONFIG.sandbox_controller_token}"
    url = (
        CONFIG.sandbox_controller_url.rstrip("/")
        + f"/runs/{CONFIG.run_root.name}/gate"
    )
    request = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            body = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"controller gate failed with HTTP {exc.code}: {detail}") from exc
    result = json.loads(body) if body else {}
    manifest = result.get("manifest")
    if not isinstance(manifest, dict):
        raise RuntimeError(f"controller gate returned no manifest: {result}")
    return manifest


@tool
def run_output_gate(export_artifacts: list[str] | None = None) -> dict[str, Any]:
    """Run the deterministic output allowlist gate for declared /outputs artifacts."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    if export_artifacts is None or not export_artifacts:
        if not DEEP_REVIEW_REQUESTS:
            return {"ok": False, "error": "no_review_request_to_export"}
        export_artifacts = list(DEEP_REVIEW_REQUESTS[-1].get("artifacts", []))
    normalized: list[str] = []
    errors: list[str] = []
    for artifact in export_artifacts:
        try:
            normalized.append(normalize_expected_artifact(artifact))
        except Exception as exc:
            errors.append(f"{artifact}: {exc}")
    if errors:
        return {"ok": False, "error": "invalid_export_artifacts", "details": errors}
    try:
        if CONFIG.sandbox_backend == "controller" and CONFIG.sandbox_controller_url:
            manifest = call_controller_gate(normalized)
        else:
            manifest = run_output_gate_artifacts(
                raw_root=CONFIG.output_dir,
                clean_root=CONFIG.clean_export_dir,
                quarantine_root=CONFIG.quarantine_dir,
                log_root=CONFIG.gate_log_dir,
                artifacts=normalized,
                run_id=CONFIG.run_root.name,
                xlsx_dangerous_formula_action=CONFIG.xlsx_dangerous_formula_action,
            )
    except Exception as exc:
        return {"ok": False, "error": f"{exc.__class__.__name__}: {exc}"}
    latest_manifest = CONFIG.runner_log_dir / "latest_gate_manifest.json"
    latest_manifest.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return {
        "ok": manifest.get("overall_status") == "pass",
        "overall_status": manifest.get("overall_status"),
        "manifest_path": "/gate_logs/gate_manifest.json",
        "clean_export_root": str(CONFIG.clean_export_dir),
        "quarantine_root": str(CONFIG.quarantine_dir),
        "artifacts": manifest.get("artifacts", []),
    }


@tool
def inspect_gate_manifest() -> dict[str, Any]:
    """Read the latest output-gate manifest."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    path = CONFIG.gate_log_dir / "gate_manifest.json"
    if not path.exists():
        return {"ok": False, "exists": False, "path": str(path)}
    try:
        return {"ok": True, "exists": True, "manifest": json.loads(path.read_text(encoding="utf-8"))}
    except Exception as exc:
        return {"ok": False, "exists": True, "error": f"{exc.__class__.__name__}: {exc}"}


@tool
def list_exported_files(root: str = "/exports", max_entries: int = 100) -> dict[str, Any]:
    """List files under the clean export area after output-gate processing."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        normalized, host_root = normalize_export_path(root) if root != "/exports" else ("/exports", CONFIG.clean_export_dir)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    max_entries = max(1, min(max_entries, 500))
    if not host_root.exists():
        return {"ok": False, "root": normalized, "exists": False, "files": []}
    files: list[dict[str, Any]] = []
    clean_root = host_root.resolve()
    for child in sorted(host_root.rglob("*"), key=lambda item: str(item).lower()):
        if len(files) >= max_entries:
            break
        resolved = child.resolve()
        if not is_relative_to(resolved, clean_root):
            continue
        rel = child.relative_to(host_root).as_posix()
        files.append(
            {
                "path": normalized.rstrip("/") + "/" + rel,
                "is_file": child.is_file(),
                "is_dir": child.is_dir(),
                "bytes": child.stat().st_size if child.is_file() else None,
                "mtime": timestamp_iso(child),
            }
        )
    return {
        "ok": True,
        "root": normalized,
        "entries_returned": len(files),
        "truncated": len(files) >= max_entries,
        "files": files,
    }


@tool
def read_exported_file(
    path: str,
    max_chars: int = 12000,
    max_rows: int = 20,
    max_cols: int = 12,
    max_pages: int = 5,
) -> dict[str, Any]:
    """Inspect a clean exported file after output-gate processing."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    try:
        normalized, host_path = normalize_export_path(path)
        max_chars = max(1000, min(max_chars, 50000))
        max_rows = max(1, min(max_rows, 100))
        max_cols = max(1, min(max_cols, 50))
        max_pages = max(1, min(max_pages, 20))
        info = inspect_file_by_type(
            host_path,
            max_chars=max_chars,
            max_rows=max_rows,
            max_cols=max_cols,
            max_pages=max_pages,
        )
        info.update({"ok": bool(info.get("exists", True)), "path": normalized, "host_path": str(host_path)})
        return info
    except Exception as exc:
        return {"ok": False, "path": path, "error": f"{exc.__class__.__name__}: {exc}"}


@tool
def inspect_exported_artifacts(max_rows: int = 20, max_cols: int = 12) -> dict[str, Any]:
    """Inspect configured expected artifacts from the clean export area."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    inspections = [
        read_exported_file.invoke(
            {"path": path, "max_rows": max_rows, "max_cols": max_cols}
        )
        for path in CONFIG.expected_artifacts
    ]
    return {
        "ok": all(item.get("ok") for item in inspections),
        "inspections": inspections,
    }


@tool
def list_quarantine_metadata() -> dict[str, Any]:
    """List quarantined raw artifacts and output-gate findings."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    manifest_path = CONFIG.gate_log_dir / "gate_manifest.json"
    manifest = None
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    quarantined = []
    for path in sorted(CONFIG.quarantine_dir.rglob("*"), key=lambda item: str(item).lower()):
        if path.is_file():
            quarantined.append(
                {
                    "path": str(path),
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path) if path.exists() else None,
                }
            )
    return {
        "ok": True,
        "quarantine_root": str(CONFIG.quarantine_dir),
        "quarantined_files": quarantined,
        "manifest_rejections": [
            item
            for item in (manifest or {}).get("artifacts", [])
            if item.get("status") == "rejected"
        ],
    }


@tool
def inspect_expected_artifacts(max_rows: int = 20, max_cols: int = 12) -> dict[str, Any]:
    """Inspect every configured expected artifact for parent review."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    inspections = [
        read_sandbox_file.invoke(
            {"path": path, "max_rows": max_rows, "max_cols": max_cols}
        )
        for path in CONFIG.expected_artifacts
    ]
    artifact_check = check_expected_artifacts(CONFIG.expected_artifacts)
    return {
        "ok": artifact_check["ok"] and all(item.get("ok") for item in inspections),
        "artifact_check": artifact_check,
        "inspections": inspections,
    }


@tool
def inspect_self_check_artifacts(max_chars: int = 16000) -> dict[str, Any]:
    """Inspect Deep Agent-generated self-check plan, script candidates, and report."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}

    output_dir = CONFIG.output_dir
    self_check_dir = output_dir / "subtasks"
    script_candidates = sorted(
        [
            path
            for path in self_check_dir.glob("self_check.*")
            if path.name not in {"self_check_plan.md", "self_check_report.md"}
            and path.is_file()
        ],
        key=lambda path: path.name.lower(),
    )
    paths = [
        "/outputs/subtasks/self_check_plan.md",
        *[
            "/outputs/" + path.relative_to(output_dir).as_posix()
            for path in script_candidates[:10]
        ],
        "/outputs/subtasks/self_check_report.md",
    ]
    inspections = [
        read_sandbox_file.invoke(
            {"path": path, "max_chars": max_chars, "max_rows": 30, "max_cols": 20}
        )
        for path in paths
    ]
    plan_exists = (self_check_dir / "self_check_plan.md").exists()
    report_exists = (self_check_dir / "self_check_report.md").exists()
    latest_policy = (
        DEEP_AGENT_EVALUATIONS[-1].get("profile", {}).get("self_check_policy", "script")
        if DEEP_AGENT_EVALUATIONS
        else "script"
    )
    script_required = latest_policy == "script"
    return {
        "ok": plan_exists and report_exists and (not script_required or bool(script_candidates)),
        "self_check_policy": latest_policy,
        "script_required": script_required,
        "plan_exists": plan_exists,
        "report_exists": report_exists,
        "script_candidates": [path.name for path in script_candidates],
        "inspections": inspections,
    }


@tool
def request_parent_review(
    artifacts: list[str],
    summary: str,
    known_issues: str = "",
) -> dict[str, Any]:
    """Ask the parent agent to review output-gate allowed artifacts before final closure."""
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    normalized_artifacts: list[str] = []
    errors: list[str] = []
    for artifact in artifacts:
        try:
            normalized_artifacts.append(normalize_expected_artifact(artifact))
        except Exception as exc:
            errors.append(f"{artifact}: {exc}")
    if errors:
        return {
            "ok": False,
            "review_requested": False,
            "error": "invalid_review_artifacts",
            "details": errors,
            "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
            "message": (
                "Review artifacts must be under /outputs and use only output-gate "
                "allowed extensions. Do not include executable self-check scripts, "
                "images, PDFs, or other helper files in request_parent_review."
            ),
        }
    preflight_findings = validate_review_artifact_preflight(normalized_artifacts)
    if preflight_findings:
        return {
            "ok": False,
            "review_requested": False,
            "error": "review_preflight_failed",
            "findings": preflight_findings,
            "message": (
                "Review artifacts are not ready. Fix the findings, rerun any "
                "required self-check, update the report, then call "
                "request_parent_review again. Do not submit placeholder or "
                "pending self-check reports."
            ),
        }
    request = {
        "attempt": len(DEEP_AGENT_EVALUATIONS) + 1,
        "artifacts": normalized_artifacts,
        "summary": summary,
        "known_issues": known_issues,
        "recorded_at": datetime.now().isoformat(timespec="seconds"),
    }
    DEEP_REVIEW_REQUESTS.append(request)
    return {
        "ok": True,
        "review_requested": True,
        "message": (
            "Parent review request recorded. Stop further work for this attempt and "
            "return a concise message listing the artifacts awaiting review."
        ),
        "request": request,
    }


def default_deep_agent_profile() -> DeepAgentProfile:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    return DeepAgentProfile(
        id="default",
        tool_name="run_deep_agent_task",
        description="Run the default sandboxed Deep Agent worker.",
        toolsets=["review", "file_read", "image_inspect", "site_crawl", "browser"],
        skill_sources=CONFIG.skill_sources,
        image=CONFIG.image,
        deep_model=CONFIG.deep_model,
        vision_model=CONFIG.vision_model,
        deep_recursion_limit=CONFIG.deep_recursion_limit,
        max_review_rounds=CONFIG.max_review_rounds,
    )


def build_deep_agent_system_prompt(profile: DeepAgentProfile) -> str:
    profile_block = ""
    if profile.system_prompt.strip():
        profile_block = (
            f"Deep Agent profile id: {profile.id}\n"
            f"Profile purpose: {profile.description}\n"
            "Profile-specific instructions:\n"
            f"{profile.system_prompt.strip()}\n\n"
        )
    toolset_lines: list[str] = []
    for toolset in profile.toolsets:
        names = list(TOOLSET_TOOL_NAMES[toolset])
        if profile.result_mode == "inline" and toolset == "review":
            names = []
        suffix = f"({', '.join(names)})" if names else "(disabled for inline result mode)"
        toolset_lines.append(f"- {toolset}: {TOOLSET_DESCRIPTIONS[toolset]} {suffix}")
    toolset_block = (
        "Available custom toolsets for this profile:\n"
        + "\n".join(toolset_lines)
        + "\n\n"
    )
    file_read_text = ""
    if "file_read" in profile.toolsets:
        image_suffix = (
            " If `image_inspect` is also available, image questions can use vision; "
            "otherwise image files return metadata only."
            if "image_inspect" in profile.toolsets
            else " Image files return metadata only because `image_inspect` is not available."
        )
        file_read_text = (
            "Use read_sandbox_file when you need a type-aware read of /input or "
            "/outputs files: it can preview text, CSV, JSON, Excel, PDFs, image "
            f"metadata, and other supported files.{image_suffix} "
        )
    image_text = ""
    if "image_inspect" in profile.toolsets:
        image_text = (
            "Use inspect_sandbox_image for focused visual review after creating "
            "crops, contact sheets, plots, or screenshots. The tool uses "
            "OpenAI image detail='original' internally for all vision reads. "
        )
    site_text = ""
    if "site_crawl" in profile.toolsets:
        site_text = (
            "For website research, do not use arbitrary search or network access "
            "from the sandbox. Instead, use the controlled site tools: "
            "crawl_allowed_site, extract_allowed_site_links, crawl_allowed_urls, "
            "search_site_crawl, read_crawled_page, and list_site_crawls. These "
            "tools enforce allowed domains, page/depth limits, response-size "
            "limits, and robots.txt before writing a local crawl index under "
            "/outputs/_site_crawl. When completeness for dated articles or "
            "collection pages matters, extract the listing links first, then "
            "crawl that explicit URL set. "
        )
    browser_text = ""
    if "browser" in profile.toolsets:
        browser_text = (
            "For interactive sites or local browser validation that require "
            "JavaScript, rendered forms, clicks, DOM checks, screenshots, or "
            "CSRF-managed browser flows, use run_playwright_task with explicit "
            "allowed_domains. Browser tasks include an egress guard: keep inputs "
            "short, search-like, and task-relevant; do not send file contents, "
            "secrets, or structured payloads through browser fields or URLs. "
            "The browser tool applies generic rate-limit avoidance by default: "
            "keep min_action_delay_ms at least 1000 unless a slower pace is "
            "needed, avoid opening many detail pages in rapid succession, and if "
            "HTTP 429 or Too Many Requests appears, back off and report the "
            "limitation rather than increasing request volume. "
        )
    if profile.result_mode == "inline":
        output_text = (
            "Return the final answer directly; use /outputs only for temporary helper "
            "files if needed. "
        )
    else:
        output_text = "Write final artifacts under /outputs. "
    if profile.input_access == "all":
        input_text = "Read task inputs only from /input. " + output_text
    elif profile.input_access == "skills_only":
        input_text = (
            "No user task input files are mounted in this profile. /input contains "
            "profile skills only; do not treat it as task evidence. " + output_text
        )
    else:
        input_text = (
            "No user task input files are mounted in this profile. " + output_text
        )
    if profile.result_mode == "inline":
        completion_contract = (
            "This profile uses inline result mode. Do not create final reviewed "
            "artifacts unless they are purely temporary helper files, and do not call "
            "request_parent_review. Return the final answer directly in your last "
            "message. Keep it concise but include the answer, confidence, evidence "
            "used, uncertainty/limitations, and a short self-check checklist. "
        )
    elif profile.self_check_policy == "script":
        completion_contract = (
            "Before requesting parent review, you must perform an autonomous self-check. "
            "This self-check is task-specific and you must design it yourself; do not "
            "wait for a prebuilt validator. Required self-check artifacts: "
            "`/outputs/subtasks/self_check_plan.md`, one executable check script such as "
            "`/outputs/subtasks/self_check.py` or `/outputs/subtasks/self_check.js`, and "
            "`/outputs/subtasks/self_check_report.md`. The plan must explain what you will verify "
            "against the user task. The script must inspect the generated artifacts and, "
            "when relevant, execute code, parse files, load workbooks, validate CSV/JSON, "
            "or run smoke tests using available local tools. If a headless browser is "
            "available for HTML tasks, use it; otherwise run the strongest available "
            "syntax/reference checks and state the limitation. Execute the self-check "
            "script with the `execute` tool. If it fails, fix the artifact and rerun the "
            "self-check before requesting review. The report must include command(s) run, "
            "pass/fail status, checked files, limitations, and remaining known issues. "
            "When describing checks for active HTML content in Markdown artifacts, escape "
            "raw tag examples such as `&lt;script&gt;`; do not write literal script tags "
            "because the output gate rejects active-content strings. Do not submit a "
            "placeholder or PENDING self-check report; request_parent_review will reject it. "
            "When you believe the expected artifacts are ready for review, you must "
            "call request_parent_review with the final artifact paths, a concise summary of "
            "what you produced, and any known issues. Include `/outputs/subtasks/self_check_plan.md` "
            "and `/outputs/subtasks/self_check_report.md` in the review request artifacts list, but "
            "do not include the executable self-check script because code is not an allowed "
            "export format. Call review after creating or updating "
            "the artifacts and completing the self-check, not before. After the review request tool returns, stop work "
            "for this attempt and respond concisely with the paths awaiting review. "
            "If this invocation contains parent correction feedback from a previous "
            "review, fix the issues first, rerun self-check, then call "
            "request_parent_review again."
        )
    elif profile.self_check_policy == "checklist":
        completion_contract = (
            "Before requesting parent review, perform a lightweight autonomous "
            "checklist self-check. Create `/outputs/subtasks/self_check_plan.md` "
            "and `/outputs/subtasks/self_check_report.md`, but do not create or "
            "execute a separate self-check script unless the task explicitly needs "
            "one. The report must state pass/fail status, checked evidence/files, "
            "limitations, and remaining known issues. When expected artifacts are "
            "ready, call request_parent_review with the final artifact paths plus "
            "`/outputs/subtasks/self_check_plan.md` and "
            "`/outputs/subtasks/self_check_report.md`. After the review request "
            "tool returns, stop work."
        )
    else:
        completion_contract = (
            "This profile does not require self-check artifacts. When expected "
            "artifacts are ready, call request_parent_review with the final artifact "
            "paths, a concise summary, and known issues. After the review request "
            "tool returns, stop work."
        )
    return (
        profile_block
        + toolset_block
        + "You are a task execution agent running in an isolated sandbox. "
        + input_text
        + f"Review/export artifacts may only use these extensions: {ALLOWED_EXPORT_EXTENSIONS_TEXT}. "
        "Do not include helper scripts, images, PDFs, or other non-allowed "
        "files in request_parent_review. "
        "Save larger scripts under /outputs before executing them. "
        "Do not use the network from the sandbox. Use installed local libraries when helpful. "
        + file_read_text
        + image_text
        + site_text
        + browser_text
        + "Cite inspected file paths and any uncertainty in your artifacts. "
        + completion_contract
    )


def deep_agent_tools_for_profile(profile: DeepAgentProfile) -> list[Any]:
    tools_by_toolset: dict[str, list[Any]] = {
        "review": [request_parent_review],
        "file_read": [read_sandbox_file],
        "image_inspect": [inspect_sandbox_image],
        "site_crawl": [
            crawl_allowed_site,
            extract_allowed_site_links,
            crawl_allowed_urls,
            search_site_crawl,
            read_crawled_page,
            list_site_crawls,
        ],
        "browser": [run_playwright_task],
    }
    selected: list[Any] = []
    seen_names: set[str] = set()
    for toolset in profile.toolsets:
        if profile.result_mode == "inline" and toolset == "review":
            continue
        for item in tools_by_toolset[toolset]:
            name = getattr(item, "name", None) or getattr(item, "__name__", str(item))
            if name in seen_names:
                continue
            selected.append(item)
            seen_names.add(name)
    return selected


def graceful_finalize_thresholds(
    recursion_limit: int,
    overrides: dict[str, int] | None = None,
) -> dict[str, int]:
    """Choose conservative finalize thresholds before LangGraph's hard limit.

    LangGraph's recursion_limit counts graph supersteps, not only visible model
    or tool calls. Deep Agents add filesystem, todo, summarization, and tool
    patching nodes, so visible message counts need to trigger finalization well
    before the hard recursion limit.
    """
    safe_limit = max(10, recursion_limit)
    finalize_model_calls = max(6, min(30, int(safe_limit * 0.16)))
    warning_model_calls = max(4, int(finalize_model_calls * 0.70))
    finalize_tool_calls = max(6, min(16, int(safe_limit * 0.12)))
    warning_tool_calls = max(4, int(finalize_tool_calls * 0.70))
    finalize_message_count = max(20, min(60, int(safe_limit * 0.38)))
    warning_message_count = max(12, int(finalize_message_count * 0.70))
    thresholds = {
        "warning_model_calls": warning_model_calls,
        "finalize_model_calls": finalize_model_calls,
        "warning_tool_calls": warning_tool_calls,
        "finalize_tool_calls": finalize_tool_calls,
        "warning_message_count": warning_message_count,
        "finalize_message_count": finalize_message_count,
        "strict_finalize_after_model_calls": 2,
    }
    for key, value in (overrides or {}).items():
        if key not in GRACEFUL_FINALIZE_CONFIG_FIELDS:
            raise ValueError(f"Unsupported graceful finalize override: {key}")
        thresholds[key] = value
    return thresholds


def execute_deep_agent_task(
    task: str,
    expected_artifacts: list[str],
    profile: DeepAgentProfile,
) -> dict[str, Any]:
    """Run one sandboxed Deep Agent attempt for allowed /outputs artifacts.

    expected_artifacts must be files under /outputs using only output-gate
    allowed extensions: .csv, .html, .json, .md, .xlsx, .yaml, .yml. For the
    generic runner, pass the configured final artifacts exactly. For multi-step
    runs, intermediate expected artifacts may be under /outputs/subtasks/. Do
    not request .py, .js, .png, .pdf, .docx, .pptx, .xlsm, or directory
    artifacts as review/export artifacts.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    effective_image = profile.image or CONFIG.image
    effective_model = profile.deep_model or CONFIG.deep_model
    effective_vision_model = profile.vision_model or CONFIG.vision_model or effective_model
    effective_recursion_limit = profile.deep_recursion_limit or CONFIG.deep_recursion_limit
    effective_max_review_rounds = profile.max_review_rounds or CONFIG.max_review_rounds
    effective_skill_sources = profile.skill_sources or []
    effective_input_dir = profile_input_dir(profile)
    try:
        effective_expected_artifacts = normalize_tool_expected_artifacts(
            expected_artifacts,
            allow_empty=profile.result_mode == "inline",
        )
        if profile.result_mode == "inline" and effective_expected_artifacts:
            return {
                "ok": False,
                "error": "inline_profile_does_not_accept_expected_artifacts",
                "message": (
                    "This profile returns an inline result. Call it with "
                    "expected_artifacts=[] and do not run output gate."
                ),
                "profile": {
                    "id": profile.id,
                    "tool_name": profile.tool_name,
                    "result_mode": profile.result_mode,
                },
            }
    except Exception as exc:
        return {
            "ok": False,
            "error": "invalid_expected_artifacts",
            "message": str(exc),
            "allowed_extensions": sorted(ALLOWED_EXTENSIONS),
        }

    attempt = len(DEEP_AGENT_EVALUATIONS) + 1
    review_start = len(DEEP_REVIEW_REQUESTS)
    if attempt > effective_max_review_rounds:
        return {
            "ok": False,
            "error": "max_review_rounds_exceeded",
            "attempt": attempt,
            "max_review_rounds": effective_max_review_rounds,
            "profile": {
                "id": profile.id,
                "tool_name": profile.tool_name,
            },
        }

    log_dir = CONFIG.runner_log_dir
    deep_prompt_path = log_dir / f"deep_agent_prompt_round_{attempt}.txt"
    deep_trace_path = log_dir / f"deep_agent_trace_round_{attempt}.json"
    cleanup_path = log_dir / f"cleanup_report_round_{attempt}.json"
    evaluation_path = log_dir / f"parent_tool_evaluation_round_{attempt}.json"
    latest_deep_prompt_path = log_dir / "deep_agent_prompt.txt"
    latest_deep_trace_path = log_dir / "deep_agent_trace.json"
    latest_cleanup_path = log_dir / "cleanup_report.json"
    latest_evaluation_path = log_dir / "parent_tool_evaluation.json"

    if profile.result_mode == "inline":
        task_with_contract = (
            "This invocation uses inline result mode.\n"
            "- Expected review/export artifacts: none.\n"
            "- Do not create final reviewed files unless they are temporary helper files.\n"
            "- Do not call request_parent_review.\n"
            "- Return the final answer directly in your last message with confidence, "
            "evidence, uncertainty/limitations, and a short self-check checklist.\n\n"
            "Task instructions:\n"
            + task
        )
    else:
        task_with_contract = (
            "Expected review/export artifacts for this invocation:\n"
            + "\n".join(f"- {path}" for path in effective_expected_artifacts)
            + "\n\nAllowed review/export extensions: "
            + ALLOWED_EXPORT_EXTENSIONS_TEXT
            + "\nDo not pass helper scripts, images, PDFs, or other non-allowed files "
            "to request_parent_review. If structured data is needed as a reviewed artifact, "
            "write CSV/XLSX/JSON/YAML or include it in a Markdown report.\n\n"
            "Task instructions:\n"
            + task
        )

    deep_prompt_path.write_text(task_with_contract, encoding="utf-8")
    latest_deep_prompt_path.write_text(task_with_contract, encoding="utf-8")

    if not configured_image_available(effective_image):
        evaluation = {
            "ok": False,
            "attempt": attempt,
            "error": "sandbox_image_missing",
            "sandbox_backend": CONFIG.sandbox_backend,
            "image": effective_image,
            "host_os": CONFIG.host_os,
            "profile": {
                "id": profile.id,
                "tool_name": profile.tool_name,
                "description": profile.description,
            },
        }
        DEEP_AGENT_EVALUATIONS.append(evaluation)
        evaluation_path.write_text(
            json.dumps(evaluation, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        latest_evaluation_path.write_text(
            json.dumps(evaluation, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return evaluation

    backend = create_configured_backend(effective_image, input_dir=effective_input_dir)
    selected_deep_tools = deep_agent_tools_for_profile(profile)
    selected_deep_tool_names = [
        getattr(item, "name", None) or getattr(item, "__name__", str(item))
        for item in selected_deep_tools
    ]
    graceful_finalize_config = graceful_finalize_thresholds(
        effective_recursion_limit,
        profile.graceful_finalize,
    )

    deep_error: dict[str, Any] | None = None
    inline_result = ""
    DEEP_AGENT_TRACE.clear()
    global ACTIVE_DEEP_AGENT_TOOLSETS, ACTIVE_DEEP_AGENT_VISION_MODEL
    try:
        ACTIVE_DEEP_AGENT_TOOLSETS = set(profile.toolsets)
        ACTIVE_DEEP_AGENT_VISION_MODEL = effective_vision_model
        deep_agent = create_deep_agent(
            model=effective_model,
            tools=selected_deep_tools,
            backend=backend,
            skills=effective_skill_sources or None,
            system_prompt=build_deep_agent_system_prompt(profile),
            middleware=[
                GracefulFinalizeMiddleware(
                    profile_id=profile.id,
                    expected_artifacts=effective_expected_artifacts,
                    result_mode=profile.result_mode,
                    self_check_policy=profile.self_check_policy,
                    **graceful_finalize_config,
                )
            ],
        )
        latest_messages: list[Any] = []
        stream_config = {
            "configurable": {
                "thread_id": f"deep-agent-{safe_tool_name(profile.id)}-task"
            },
            "recursion_limit": effective_recursion_limit,
        }
        for chunk in deep_agent.stream(
            {"messages": [{"role": "user", "content": task_with_contract}]},
            config=stream_config,
            stream_mode="values",
        ):
            if isinstance(chunk, dict) and chunk.get("messages"):
                latest_messages = list(chunk["messages"])
                DEEP_AGENT_TRACE[:] = trace_messages(latest_messages, content_limit=2600)
                latest_deep_trace_path.write_text(
                    json.dumps(DEEP_AGENT_TRACE, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
        if latest_messages:
            inline_result = content_text(getattr(latest_messages[-1], "content", ""))
            DEEP_AGENT_TRACE[:] = trace_messages(latest_messages, content_limit=2600)
        deep_trace_path.write_text(
            json.dumps(DEEP_AGENT_TRACE, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        latest_deep_trace_path.write_text(
            json.dumps(DEEP_AGENT_TRACE, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as exc:
        deep_error = {"type": exc.__class__.__name__, "message": str(exc)[:2000]}
        error_trace: dict[str, Any] | list[dict[str, Any]]
        error_trace = (
            {"error": deep_error, "partial_trace": DEEP_AGENT_TRACE}
            if DEEP_AGENT_TRACE
            else {"error": deep_error}
        )
        deep_trace_path.write_text(
            json.dumps(error_trace, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        latest_deep_trace_path.write_text(
            json.dumps(error_trace, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    finally:
        ACTIVE_DEEP_AGENT_TOOLSETS = set()
        ACTIVE_DEEP_AGENT_VISION_MODEL = ""
        workspace_dir = backend.workspace_dir
        backend.cleanup()
        workspace_exists = workspace_dir.exists()
        cleanup_ok = (
            not workspace_exists
            if CONFIG.sandbox_backend == "podman"
            else True
        )
        cleanup_payload = {
            "sandbox_id": backend.id,
            "workspace_dir": str(workspace_dir),
            "workspace_exists_after_cleanup": workspace_exists,
            "cleanup_policy": CONFIG.sandbox_backend,
            "cleanup_ok": cleanup_ok,
            "profile_id": profile.id,
        }
        cleanup_path.write_text(
            json.dumps(cleanup_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        latest_cleanup_path.write_text(
            json.dumps(cleanup_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    artifact_check = (
        {"artifacts": [], "ok": True, "missing": []}
        if profile.result_mode == "inline"
        else check_expected_artifacts(effective_expected_artifacts)
    )
    cleanup = json.loads(cleanup_path.read_text(encoding="utf-8"))
    attempt_review_requests = DEEP_REVIEW_REQUESTS[review_start:]
    self_check_dir = CONFIG.output_dir / "subtasks"
    self_check_scripts = sorted(
        [
            path.relative_to(CONFIG.output_dir).as_posix()
            for path in self_check_dir.glob("self_check.*")
            if path.name not in {"self_check_plan.md", "self_check_report.md"}
            and path.is_file()
        ]
    )
    evaluation = {
        "attempt": attempt,
        "max_review_rounds": effective_max_review_rounds,
        "profile": {
            "id": profile.id,
            "tool_name": profile.tool_name,
            "description": profile.description,
            "toolsets": profile.toolsets,
            "result_mode": profile.result_mode,
            "self_check_policy": profile.self_check_policy,
            "available_tools": selected_deep_tool_names,
            "graceful_finalize": graceful_finalize_config,
            "image": effective_image,
            "deep_model": effective_model,
            "vision_model": effective_vision_model,
            "deep_recursion_limit": effective_recursion_limit,
            "skill_sources": effective_skill_sources,
            "input_access": profile.input_access,
        },
        "runtime": {
            "host_os": CONFIG.host_os,
            "sandbox_backend": CONFIG.sandbox_backend,
            "image": effective_image,
            "podman_bin": CONFIG.podman_bin,
            "wsl_distro": CONFIG.wsl_distro if CONFIG.host_os == "windows" else None,
            "selinux_relabel": CONFIG.selinux_relabel,
            "egress_proxy_enabled": bool(CONFIG.egress_proxy_url),
        },
        "expected_artifacts": effective_expected_artifacts,
        "configured_final_artifacts": CONFIG.expected_artifacts,
        "parent_provided_expected_artifacts": expected_artifacts,
        "artifact_check": artifact_check,
        "deep_error": deep_error,
        "cleanup": cleanup,
        "review_requested": bool(attempt_review_requests),
        "review_requests": attempt_review_requests,
        "inline_result": inline_result if profile.result_mode == "inline" else "",
        "self_check": {
            "plan_exists": (self_check_dir / "self_check_plan.md").exists(),
            "report_exists": (self_check_dir / "self_check_report.md").exists(),
            "script_candidates": self_check_scripts,
        },
        "salvage": build_unreviewed_artifact_salvage(
            profile=profile,
            effective_expected_artifacts=effective_expected_artifacts,
            artifact_check=artifact_check,
            deep_error=deep_error,
            review_requested=bool(attempt_review_requests),
            self_check_dir=self_check_dir,
        ),
        "deep_tool_calls": [
            name for item in DEEP_AGENT_TRACE for name in item.get("tool_calls", [])
        ],
        "skill_usage": extract_skill_usage_from_trace(
            DEEP_AGENT_TRACE,
            effective_skill_sources,
        ),
    }
    evaluation["ok"] = (
        artifact_check["ok"]
        and deep_error is None
        and cleanup.get("cleanup_ok") is True
        and (profile.result_mode == "inline" or bool(attempt_review_requests))
        and (profile.result_mode != "inline" or bool(inline_result.strip()))
    )
    evaluation_path.write_text(
        json.dumps(evaluation, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    latest_evaluation_path.write_text(
        json.dumps(evaluation, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    DEEP_AGENT_EVALUATIONS.append(evaluation)
    return evaluation


@tool
def run_deep_agent_task(task: str, expected_artifacts: list[str]) -> dict[str, Any]:
    """Run the default sandboxed Deep Agent attempt for allowed /outputs artifacts.

    expected_artifacts must be files under /outputs using only output-gate
    allowed extensions: .csv, .html, .json, .md, .xlsx, .yaml, .yml. For the
    generic runner, pass the configured final artifacts exactly. For multi-step
    runs, intermediate expected artifacts may be under /outputs/subtasks/. Do
    not request .py, .js, .png, .pdf, .docx, .pptx, .xlsm, or directory
    artifacts as review/export artifacts.
    """
    if CONFIG is None:
        return {"ok": False, "error": "runner_config_missing"}
    return execute_deep_agent_task(task, expected_artifacts, default_deep_agent_profile())


def make_deep_agent_profile_tool(profile: DeepAgentProfile) -> Any:
    def run_profile_deep_agent(
        task: str,
        expected_artifacts: list[str],
    ) -> dict[str, Any]:
        return execute_deep_agent_task(task, expected_artifacts, profile)

    run_profile_deep_agent.__name__ = profile.tool_name
    available_tools = ", ".join(tool_names_for_profile(profile)) or "none"
    if profile.result_mode == "inline":
        mode_contract = (
            "This profile returns an inline result. Call it with "
            "expected_artifacts=[]; it will not request parent review and the "
            "parent must not run output gate for this call. Inspect the returned "
            "inline_result and retry only if it materially fails the task."
        )
    else:
        mode_contract = (
            "This profile returns reviewed artifacts. expected_artifacts must be "
            "files under /outputs using only output-gate allowed extensions: "
            ".csv, .html, .json, .md, .xlsx, .yaml, .yml. Run output gate only "
            "after the tool reports review_requested=true."
        )
    run_profile_deep_agent.__doc__ = (
        f"{profile.description}\n\n"
        f"Available toolsets: {', '.join(profile.toolsets)}. "
        f"Available custom tools: {available_tools}. "
        f"result_mode={profile.result_mode}; "
        f"self_check_policy={profile.self_check_policy}.\n\n"
        f"{mode_contract} "
        "Do not request .py, .js, .png, .pdf, .docx, .pptx, .xlsm, or directory "
        "artifacts as review/export artifacts."
    )
    tool_description = (
        f"{profile.description} Available toolsets: {', '.join(profile.toolsets)}. "
        f"Available custom tools: {available_tools}. "
        f"result_mode={profile.result_mode}; self_check_policy={profile.self_check_policy}."
    )
    return tool(profile.tool_name, description=tool_description)(run_profile_deep_agent)


def active_deep_agent_tools() -> list[Any]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    if CONFIG.deep_agent_profiles:
        return [make_deep_agent_profile_tool(profile) for profile in CONFIG.deep_agent_profiles]
    return [run_deep_agent_task]


def active_deep_agent_tool_instruction() -> str:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")
    if CONFIG.deep_agent_profiles:
        names = ", ".join(profile.tool_name for profile in CONFIG.deep_agent_profiles)
        return f"the most appropriate Deep Agent profile tool ({names})"
    return "run_deep_agent_task"


def run_parent_agent() -> dict[str, Any]:
    if CONFIG is None:
        raise RuntimeError("Runner config is not initialized.")

    parent_prompt = build_parent_prompt()
    deep_tool_instruction = active_deep_agent_tool_instruction()
    if CONFIG.deep_agent_profiles:
        attempts_instruction = (
            f"Profile tools may define their own max_review_rounds; runner default is "
            f"{CONFIG.max_review_rounds}. Never call a Deep Agent tool after it reports "
            "max_review_rounds_exceeded."
        )
    else:
        attempts_instruction = (
            f"The maximum Deep Agent attempts is {CONFIG.max_review_rounds}. Never call "
            "run_deep_agent_task after the tool reports max_review_rounds_exceeded."
        )
    (CONFIG.runner_log_dir / "parent_prompt.txt").write_text(parent_prompt, encoding="utf-8")
    (CONFIG.runner_log_dir / "input_manifest.json").write_text(
        json.dumps(input_manifest(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    parent_tools = [
        *active_deep_agent_tools(),
        run_output_gate,
        inspect_gate_manifest,
        list_exported_files,
        read_exported_file,
        inspect_exported_artifacts,
        list_quarantine_metadata,
    ]
    if CONFIG.allow_raw_parent_inspection:
        parent_tools.extend(
            [
                list_sandbox_files,
                read_sandbox_file,
                inspect_sandbox_file,
                inspect_sandbox_image,
                inspect_expected_artifacts,
                inspect_self_check_artifacts,
                crawl_allowed_site,
                extract_allowed_site_links,
                crawl_allowed_urls,
                search_site_crawl,
                read_crawled_page,
                list_site_crawls,
            ]
        )

    parent_agent = create_agent(
        model=CONFIG.parent_model,
        tools=parent_tools,
        system_prompt=(
            "You are a parent HITL reviewer and orchestrator. The Deep Agent does the "
            "implementation. Do not edit files yourself and do not perform the "
            f"implementation yourself. Required workflow: (1) choose and call {deep_tool_instruction}. "
            "The selected tool description states its result_mode. For result_mode=inline, "
            "pass expected_artifacts=[], do not run output gate, inspect the returned "
            "inline_result directly, and close when ok=true and the inline result materially "
            "answers the user task; retry with concise correction instructions only if a "
            "material issue remains and attempts remain. For result_mode=artifact, pass the "
            f"configured final artifacts as expected_artifacts using only output-gate allowed "
            f"paths ({ALLOWED_EXPORT_EXTENSIONS_TEXT}); do not add root-level self-check files "
            "to expected_artifacts. (2) For artifact-result calls, check whether the result "
            "has review_requested=true. If review was requested, call run_output_gate with "
            "no explicit artifact override so the latest request_parent_review artifact list "
            "is gated, including any self-check plan/report files, before reading produced "
            "files. (3) Inspect the gate manifest and only read clean exports with "
            "inspect_exported_artifacts, read_exported_file, or list_exported_files. Do not "
            "read raw /outputs in production mode. If the gate rejects files, use "
            "inspect_gate_manifest and list_quarantine_metadata to cite concrete findings, "
            "then ask the Deep Agent to repair them. (4) If required artifact files are "
            "missing from clean exports, required self-check plan/report files are missing "
            "for a profile that uses script/checklist self_check_policy, gate failures were "
            "ignored, or the inspected clean artifact materially fails the user task, and "
            f"at least one Deep Agent attempt remains, call {deep_tool_instruction} again with "
            "concise correction instructions that include your findings. (5) Run output gate "
            "and inspect again after every artifact-result review request, and close only "
            "when no material issues remain or no attempts remain. If an artifact-result "
            "Deep Agent does not request review, treat that as a material protocol issue; "
            "if the tool result includes salvage.available=true, do not run output gate "
            "on those unreviewed files, but use the salvage summary to ask the next "
            "Deep Agent attempt to reuse the existing /outputs candidates, finish or "
            "repair self-check artifacts, and call request_parent_review promptly. "
            f"if attempts remain, call {deep_tool_instruction} again instructing it to produce/update "
            "artifacts, run the profile-appropriate self-check, and call request_parent_review. "
            f"{attempts_instruction} "
            "In the final response, report attempts used, files inspected, material findings, "
            "gate status when applicable, self-check status when applicable, inline_result "
            "status when applicable, remaining issues if any, whether the last Deep Agent "
            "attempt requested review for artifact-result profiles, and clean export paths "
            "when applicable. Keep claims tied to gate manifest, inspected clean-export "
            "evidence, or the returned inline_result depending on the selected profile."
        ),
    )
    parent_result = parent_agent.invoke(
        {"messages": [{"role": "user", "content": parent_prompt}]},
        config={
            "configurable": {"thread_id": "parent-agent-generic-task"},
            "recursion_limit": CONFIG.parent_recursion_limit,
        },
    )
    messages = parent_result["messages"]
    parent_trace = trace_messages(messages, content_limit=2600)
    (CONFIG.runner_log_dir / "parent_agent_trace.json").write_text(
        json.dumps(parent_trace, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    final_text = content_text(getattr(messages[-1], "content", ""))
    parent_tool_calls = [name for item in parent_trace for name in item.get("tool_calls", [])]
    evaluation_path = CONFIG.runner_log_dir / "parent_tool_evaluation.json"
    evaluation = json.loads(evaluation_path.read_text(encoding="utf-8")) if evaluation_path.exists() else {}
    review_rounds = list(DEEP_AGENT_EVALUATIONS)
    report = [
        "# Generic Parent Agent Run",
        "",
        "## Inputs",
        "",
        "```json",
        json.dumps(input_manifest(), ensure_ascii=False, indent=2),
        "```",
        "",
        "## Expected Artifacts",
        "",
        "```json",
        json.dumps(CONFIG.expected_artifacts, ensure_ascii=False, indent=2),
        "```",
        "",
        "## Deep Agent Profiles",
        "",
        "```json",
        json.dumps(
            [
                {
                    "id": profile.id,
                    "tool_name": profile.tool_name,
                    "description": profile.description,
                    "toolsets": profile.toolsets,
                    "input_access": profile.input_access,
                    "result_mode": profile.result_mode,
                    "self_check_policy": profile.self_check_policy,
                    "graceful_finalize": profile.graceful_finalize,
                    "image": profile.image or CONFIG.image,
                    "deep_model": profile.deep_model or CONFIG.deep_model,
                    "vision_model": (
                        profile.vision_model
                        or CONFIG.vision_model
                        or profile.deep_model
                        or CONFIG.deep_model
                    ),
                    "skill_sources": profile.skill_sources,
                }
                for profile in CONFIG.deep_agent_profiles
            ]
            or [
                {
                    "id": "default",
                    "tool_name": "run_deep_agent_task",
                    "description": "Default sandboxed Deep Agent worker.",
                    "toolsets": default_deep_agent_profile().toolsets,
                    "input_access": "all",
                    "result_mode": "artifact",
                    "self_check_policy": "script",
                    "image": CONFIG.image,
                    "deep_model": CONFIG.deep_model,
                    "vision_model": CONFIG.vision_model or CONFIG.deep_model,
                    "skill_sources": CONFIG.skill_sources,
                }
            ],
            ensure_ascii=False,
            indent=2,
        ),
        "```",
        "",
        "## Parent Tool Calls",
        "",
        "```json",
        json.dumps(parent_tool_calls, ensure_ascii=False, indent=2),
        "```",
        "",
        "## Parent Final Response",
        "",
        final_text,
        "",
        "## Latest Generic Artifact Evaluation",
        "",
        "```json",
        json.dumps(evaluation, ensure_ascii=False, indent=2),
        "```",
        "",
        "## Deep Agent Attempt Evaluations",
        "",
        "```json",
        json.dumps(review_rounds, ensure_ascii=False, indent=2),
        "```",
        "",
    ]
    (CONFIG.runner_log_dir / "parent_agent_report.md").write_text(
        "\n".join(report),
        encoding="utf-8",
    )
    return {
        "parent_tool_calls": parent_tool_calls,
        "final_text": final_text,
        "evaluation": evaluation,
        "deep_agent_attempts": review_rounds,
        "output_dir": str(CONFIG.run_root),
        "raw_output_dir": str(CONFIG.output_dir),
        "clean_export_dir": str(CONFIG.clean_export_dir),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generic parent-agent runner: stage inputs, delegate one task to a Deep Agent "
            "through a tool, inspect reviewed artifacts or inline results, and optionally "
            "ask for bounded repairs."
        )
    )
    prompt_group = parser.add_mutually_exclusive_group(required=True)
    prompt_group.add_argument("--prompt-file", help="UTF-8 text file containing the task prompt.")
    prompt_group.add_argument("--prompt", help="Task prompt text.")
    parser.add_argument(
        "--input",
        action="append",
        default=[],
        help="Input mapping. Use HOST_PATH=/input/name or just HOST_PATH to use the basename.",
    )
    parser.add_argument(
        "--expected-artifact",
        action="append",
        default=[],
        help="Expected output path under /outputs. Repeat for multiple artifacts.",
    )
    parser.add_argument(
        "--skill-source",
        action="append",
        default=[],
        help=(
            "Skill source directory to expose to Deep Agent. Use HOST_DIR=/input/skills "
            "or just HOST_DIR to stage it at /input/skills. The source should contain "
            "skill-name/SKILL.md directories."
        ),
    )
    parser.add_argument(
        "--deep-agent-profile",
        action="append",
        default=[],
        help="Deep Agent profile YAML/JSON file. Repeat to expose multiple profile tools.",
    )
    parser.add_argument(
        "--deep-agent-profile-dir",
        action="append",
        default=[],
        help="Directory containing Deep Agent profile .json/.yaml/.yml files.",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Host output directory under this workspace's outputs/ directory.",
    )
    parser.add_argument("--image", default="localhost/python-data-sandbox:latest")
    parser.add_argument(
        "--sandbox-backend",
        choices=["podman", "controller"],
        default=os.getenv("SANDBOX_BACKEND", "podman"),
        help="Sandbox backend. Use controller inside Docker Compose; use podman for direct host runs.",
    )
    parser.add_argument(
        "--sandbox-controller-url",
        default=os.getenv("SANDBOX_CONTROLLER_URL", ""),
        help="sandbox-controller base URL for --sandbox-backend controller.",
    )
    parser.add_argument(
        "--sandbox-controller-token",
        default=os.getenv("SANDBOX_CONTROLLER_TOKEN", ""),
        help="Bearer token for sandbox-controller, if configured.",
    )
    parser.add_argument(
        "--egress-proxy-url",
        default=os.getenv("EGRESS_PROXY_URL", ""),
        help=(
            "Optional HTTP proxy URL used by network-enabled tools. When set, "
            "browser/crawler tools receive a signed per-call allowlist token."
        ),
    )
    parser.add_argument(
        "--egress-proxy-signing-secret",
        default=os.getenv("EGRESS_PROXY_SIGNING_SECRET", ""),
        help="Shared signing secret for per-tool egress proxy tokens.",
    )
    parser.add_argument(
        "--host-os",
        choices=["auto", "windows", "linux"],
        default="auto",
        help="Host OS/runtime mode. windows uses WSL Podman; linux uses native Podman.",
    )
    parser.add_argument("--wsl-distro", default="Ubuntu-22.04")
    parser.add_argument(
        "--wsl-no-sudo",
        action="store_true",
        help="On --host-os windows, call WSL podman without sudo.",
    )
    parser.add_argument(
        "--podman-bin",
        default="podman",
        help="Podman executable for --host-os linux.",
    )
    parser.add_argument(
        "--selinux-relabel",
        action="store_true",
        help="On --host-os linux, add SELinux relabeling to bind mounts.",
    )
    parser.add_argument("--parent-model", default="openai:gpt-5.2")
    parser.add_argument("--deep-model", default="openai:gpt-5.2")
    parser.add_argument(
        "--vision-model",
        default="",
        help=(
            "Optional model for inspect_sandbox_image/read_sandbox_file image vision. "
            "Defaults to the active Deep Agent model. A profile-level vision_model "
            "overrides this value."
        ),
    )
    parser.add_argument(
        "--tile-small-images-for-vision",
        action="store_true",
        help=(
            "Experimental: when inspect_sandbox_image sends a small image to vision, "
            "tile it into a repeated grid first. Disabled by default."
        ),
    )
    parser.add_argument(
        "--vision-tile-max-side",
        type=int,
        default=128,
        help="Maximum image side in pixels for --tile-small-images-for-vision.",
    )
    parser.add_argument(
        "--vision-tile-grid",
        type=int,
        default=5,
        help="Grid size for --tile-small-images-for-vision. Default 5 makes 25 copies.",
    )
    parser.add_argument("--parent-recursion-limit", type=int, default=12)
    parser.add_argument("--deep-recursion-limit", type=int, default=80)
    parser.add_argument(
        "--max-review-rounds",
        type=int,
        default=2,
        help="Maximum number of times the parent may call the Deep Agent tool.",
    )
    parser.add_argument(
        "--no-clear-output",
        action="store_true",
        help="Do not clear the output directory before running. Not recommended for evaluations.",
    )
    parser.add_argument(
        "--keep-staged-input",
        action="store_true",
        help="Keep the temporary staged input directory under work/ after the run.",
    )
    parser.add_argument(
        "--allow-raw-parent-inspection",
        action="store_true",
        help="Development-only: expose raw /outputs inspection tools to the parent agent.",
    )
    parser.add_argument(
        "--xlsx-dangerous-formula-action",
        choices=["reject", "stringify"],
        default="reject",
        help="Output-gate action for dangerous xlsx formulas. Default rejects for repair.",
    )
    return parser.parse_args()


def configure_console_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            try:
                reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                pass


def main() -> None:
    global CONFIG
    configure_console_encoding()
    args = parse_args()

    load_dotenv(ROOT / ".env.local", override=False)
    load_env_local(ROOT / ".env.local")
    if not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError("OPENAI_API_KEY is missing.")

    prompt = (
        Path(args.prompt_file).read_text(encoding="utf-8")
        if args.prompt_file
        else args.prompt
    )
    output_arg = Path(args.output_dir)
    output_base = (
        Path(os.getenv("RUNS_ROOT", "/srv/sandbox-tool/runs"))
        if args.sandbox_backend == "controller"
        else ROOT
    )
    run_root = require_safe_output_dir(
        output_arg.resolve() if output_arg.is_absolute() else (output_base / output_arg).resolve(),
        sandbox_backend=args.sandbox_backend,
    )
    if not args.no_clear_output:
        clear_directory_contents(run_root)
    run_dirs = prepare_run_directories(run_root)

    input_dir = run_dirs["input_dir"]
    input_mappings = stage_inputs(args.input, input_dir)
    skill_sources = stage_skill_sources(args.skill_source, input_dir)
    deep_agent_profiles = load_deep_agent_profiles(
        args.deep_agent_profile,
        args.deep_agent_profile_dir,
    )
    materialize_deep_agent_profiles(deep_agent_profiles, input_dir, skill_sources)
    expected_artifacts = [normalize_expected_artifact(path) for path in args.expected_artifact]
    host_os = resolve_host_os(args.host_os)

    CONFIG = RunnerConfig(
        prompt=prompt,
        run_root=run_dirs["run_root"],
        output_dir=run_dirs["raw_output_dir"],
        clean_export_dir=run_dirs["clean_export_dir"],
        quarantine_dir=run_dirs["quarantine_dir"],
        gate_log_dir=run_dirs["gate_log_dir"],
        runner_log_dir=run_dirs["runner_log_dir"],
        input_dir=input_dir,
        workspace_dir=run_dirs["workspace_dir"],
        input_mappings=input_mappings,
        expected_artifacts=expected_artifacts,
        skill_sources=skill_sources,
        image=args.image,
        wsl_distro=args.wsl_distro,
        parent_model=args.parent_model,
        deep_model=args.deep_model,
        vision_model=args.vision_model,
        parent_recursion_limit=args.parent_recursion_limit,
        deep_recursion_limit=args.deep_recursion_limit,
        max_review_rounds=max(1, args.max_review_rounds),
        host_os=host_os,
        podman_bin=args.podman_bin,
        wsl_use_sudo=not args.wsl_no_sudo,
        selinux_relabel=args.selinux_relabel,
        clear_output=not args.no_clear_output,
        keep_staged_input=args.keep_staged_input,
        allow_raw_parent_inspection=args.allow_raw_parent_inspection,
        sandbox_backend=args.sandbox_backend,
        sandbox_controller_url=args.sandbox_controller_url,
        sandbox_controller_token=args.sandbox_controller_token,
        egress_proxy_url=args.egress_proxy_url,
        egress_proxy_signing_secret=args.egress_proxy_signing_secret,
        xlsx_dangerous_formula_action=args.xlsx_dangerous_formula_action,
        tile_small_images_for_vision=args.tile_small_images_for_vision,
        vision_tile_max_side=args.vision_tile_max_side,
        vision_tile_grid=args.vision_tile_grid,
        deep_agent_profiles=deep_agent_profiles,
    )

    try:
        result = run_parent_agent()
    finally:
        if CONFIG is not None and not CONFIG.keep_staged_input and CONFIG.input_dir.exists():
            shutil.rmtree(CONFIG.input_dir, ignore_errors=True)

    print("Parent tool calls:", result["parent_tool_calls"])
    print("Deep Agent attempts:", len(result["deep_agent_attempts"]))
    print("Evaluation ok:", result["evaluation"].get("ok"))
    print("Output dir:", result["output_dir"])
    print("Final response:")
    print(result["final_text"])


if __name__ == "__main__":
    main()
